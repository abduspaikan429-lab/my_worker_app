# modules/purlins_summary.py
"""
檩条下料单自动合计与规格汇总模块
====================================
功能：
    1. 读取多 Sheet 的檩条下料单 Excel
    2. 对每个子表自动定位数量/长度/规格列，扫描有效数据行
    3. 清除旧合计行，写入三种规格（220*100*3/4/5）SUMIF 公式合计行 + 总合计行
    4. 生成"汇总统计" Sheet（各子表合计 + 按规格统计两块）
    5. 返回处理后的 Excel bytes 供 Streamlit 下载
"""

from io import BytesIO
from collections import defaultdict

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


# ── 目标规格常量 ──────────────────────────────────────────────────────────────
SPEC_ORDER = ["220*100*3", "220*100*4", "220*100*5"]
TARGET_SPECS = set(SPEC_ORDER)


def safe_to_num(value):
    """将单元格值安全转换为 float，无法转换时返回 None。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        s = value.strip()
        if s == '':
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def add_total_and_summary(file_input) -> tuple[bytes, dict, str | None]:
    """
    读取上传的檩条下料单 Excel，为每个子表写入规格合计行与总合计行，
    并在首位生成"汇总统计" Sheet。

    Parameters
    ----------
    file_input : file-like object（BytesIO 或 Streamlit UploadedFile）
        上传的源 Excel 文件流

    Returns
    -------
    (excel_bytes, stats, error)
        excel_bytes : bytes | None    处理后的 Excel 二进制内容
        stats       : dict            统计信息（供 UI 展示）
        error       : str | None      出错时的错误信息
    """
    try:
        wb = openpyxl.load_workbook(file_input)
    except Exception as e:
        return None, {}, f"无法读取 Excel 文件：{e}"

    sheet_info_list = []               # 每个子表的关键信息（供汇总页引用）
    spec_stats = defaultdict(lambda: [0, 0])   # spec → [总根数, 总长度mm]
    processed_sheets = 0

    sheets_to_process = [ws for ws in wb.worksheets if ws.title != "汇总统计"]

    if not sheets_to_process:
        return None, {}, "未找到可处理的工作表（所有 Sheet 均名为'汇总统计'或文件为空）。"

    for ws in sheets_to_process:
        # ── 1. 定位列标题 ────────────────────────────────────────────────────
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val is not None:
                headers[str(val).strip()] = col

        qty_col  = headers.get("数量（根）")
        len_col  = headers.get("长度（mm）")
        spec_col = headers.get("规格（mm）")
        serial_col = 1

        if not qty_col or not spec_col:
            continue   # 缺少关键列，跳过该 Sheet

        # ── 2. 扫描数据行 ────────────────────────────────────────────────────
        total_qty = 0
        total_len = 0
        valid_rows = 0
        data_start_row = 0
        last_data_row = 0

        for row in range(2, ws.max_row + 1):
            serial_val = ws.cell(row=row, column=serial_col).value
            if safe_to_num(serial_val) is None:
                continue
            if valid_rows == 0:
                data_start_row = row

            spec_val = ws.cell(row=row, column=spec_col).value
            spec_str = str(spec_val).strip() if spec_val is not None else ""
            if spec_str == "":
                continue

            qty_val = ws.cell(row=row, column=qty_col).value
            qty_num = safe_to_num(qty_val)
            if qty_num is None or qty_num <= 0:
                continue

            valid_rows += 1
            last_data_row = row
            total_qty += qty_num

            row_len = 0
            if len_col:
                len_val = ws.cell(row=row, column=len_col).value
                len_num = safe_to_num(len_val)
                if len_num is not None:
                    row_len = len_num
                    total_len += row_len

            if spec_str in TARGET_SPECS:
                spec_stats[spec_str][0] += qty_num
                spec_stats[spec_str][1] += row_len

        if valid_rows == 0:
            continue

        # ── 3. 清除旧合计行 ──────────────────────────────────────────────────
        rows_to_delete = []
        for r in range(last_data_row + 1, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=1).value
            if cell_val is not None and "合计" in str(cell_val):
                rows_to_delete.append(r)
        for r in reversed(rows_to_delete):
            ws.delete_rows(r, 1)

        # ── 4. 写入 SUMIF 规格合计行 + 总合计行 ─────────────────────────────
        bold = Font(bold=True)
        qty_letter  = get_column_letter(qty_col)
        spec_letter = get_column_letter(spec_col)
        qty_range   = f"{qty_letter}{data_start_row}:{qty_letter}{last_data_row}"
        spec_range  = f"{spec_letter}{data_start_row}:{spec_letter}{last_data_row}"
        len_range = len_letter = None
        if len_col:
            len_letter = get_column_letter(len_col)
            len_range  = f"{len_letter}{data_start_row}:{len_letter}{last_data_row}"

        # 三种规格合计行（last_data_row+1 / +2 / +3）
        for i, spec in enumerate(SPEC_ORDER):
            r = last_data_row + 1 + i
            ws.cell(row=r, column=1, value=f"{spec} 合计").font = bold
            ws.cell(row=r, column=qty_col,
                    value=f'=SUMIF({spec_range},"{spec}",{qty_range})').font = bold
            if len_col and len_range:
                ws.cell(row=r, column=len_col,
                        value=f'=SUMIF({spec_range},"{spec}",{len_range})').font = bold

        # 总合计行（last_data_row+4）
        total_row = last_data_row + 4
        ws.cell(row=total_row, column=1, value="合计").font = bold
        ws.cell(row=total_row, column=qty_col,
                value=f"=SUM({qty_range})").font = bold
        if len_col and len_range:
            ws.cell(row=total_row, column=len_col,
                    value=f"=SUM({len_range})").font = bold

        # ── 5. 记录子表信息 ──────────────────────────────────────────────────
        sheet_info_list.append({
            'title': ws.title,
            'qty_col': qty_col,
            'len_col': len_col,
            'last_data_row': last_data_row,
            'total_qty': total_qty,
            'total_len': total_len,
        })
        processed_sheets += 1

    if not sheet_info_list:
        return None, {}, "所有工作表均缺少【数量（根）】或【规格（mm）】列，无法处理。请检查列标题是否正确。"

    # ══════════════════════════════════════════════════════════════════════════
    # 汇总统计页
    # ══════════════════════════════════════════════════════════════════════════
    if "汇总统计" in wb.sheetnames:
        wb.remove(wb["汇总统计"])
    ws_sum = wb.create_sheet("汇总统计", 0)

    thin    = Border(left=Side('thin'), right=Side('thin'),
                     top=Side('thin'), bottom=Side('thin'))
    h_font  = Font(bold=True, size=11)
    n_font  = Font(size=11)

    # ── 第一块：各子表合计 ────────────────────────────────────────────────────
    ws_sum.merge_cells("A1:C1")
    ws_sum["A1"] = "各子表合计"
    ws_sum["A1"].font = Font(bold=True, size=13)
    ws_sum.cell(2, 1, "工作表名称").font    = h_font
    ws_sum.cell(2, 2, "数量（根）合计").font = h_font
    ws_sum.cell(2, 3, "长度（mm）合计").font = h_font
    for c in range(1, 4):
        ws_sum.cell(2, c).border = thin

    data_start = 3
    for idx, info in enumerate(sheet_info_list):
        cur = data_start + idx
        ws_sum.cell(cur, 1, info['title']).font = n_font

        qty_ref = f"'{info['title']}'!{get_column_letter(info['qty_col'])}{info['last_data_row'] + 4}"
        ws_sum.cell(cur, 2, f"=SUM({qty_ref})").font = n_font

        if info['len_col']:
            len_ref = f"'{info['title']}'!{get_column_letter(info['len_col'])}{info['last_data_row'] + 4}"
            ws_sum.cell(cur, 3, f"=SUM({len_ref})").font = n_font
        else:
            ws_sum.cell(cur, 3, 0).font = n_font

        for c in range(1, 4):
            ws_sum.cell(cur, c).border = thin

    # 所有子表合计行
    sum_total_row = data_start + len(sheet_info_list)
    ws_sum.cell(sum_total_row, 1, "所有子表合计").font = Font(bold=True)
    ws_sum.cell(sum_total_row, 2, f"=SUM(B{data_start}:B{sum_total_row-1})").font = Font(bold=True)
    ws_sum.cell(sum_total_row, 3, f"=SUM(C{data_start}:C{sum_total_row-1})").font = Font(bold=True)
    for c in range(1, 4):
        ws_sum.cell(sum_total_row, c).border = thin

    # ── 第二块：按规格统计 ────────────────────────────────────────────────────
    split_row = sum_total_row + 2
    ws_sum.merge_cells(start_row=split_row, start_column=1,
                       end_row=split_row, end_column=3)
    ws_sum.cell(split_row, 1, "按规格统计（220*100*3/4/5）").font = Font(bold=True, size=13)

    spec_h = split_row + 1
    ws_sum.cell(spec_h, 1, "规格（mm）").font    = h_font
    ws_sum.cell(spec_h, 2, "数量（根）合计").font = h_font
    ws_sum.cell(spec_h, 3, "长度（mm）合计").font = h_font
    for c in range(1, 4):
        ws_sum.cell(spec_h, c).border = thin

    for spec_idx, spec in enumerate(SPEC_ORDER):
        r = spec_h + 1 + spec_idx
        ws_sum.cell(r, 1, spec).font = n_font

        qty_parts, len_parts = [], []
        for info in sheet_info_list:
            qty_cell = f"'{info['title']}'!{get_column_letter(info['qty_col'])}{info['last_data_row'] + 1 + spec_idx}"
            qty_parts.append(qty_cell)
            if info['len_col']:
                len_cell = f"'{info['title']}'!{get_column_letter(info['len_col'])}{info['last_data_row'] + 1 + spec_idx}"
                len_parts.append(len_cell)

        ws_sum.cell(r, 2, "=SUM(" + ",".join(qty_parts) + ")" if qty_parts else 0).font = n_font
        ws_sum.cell(r, 3, "=SUM(" + ",".join(len_parts) + ")" if len_parts else 0).font = n_font
        for c in range(1, 4):
            ws_sum.cell(r, c).border = thin

    # 列宽
    ws_sum.column_dimensions['A'].width = 22
    ws_sum.column_dimensions['B'].width = 18
    ws_sum.column_dimensions['C'].width = 18

    # ── 输出 bytes ────────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)

    # 汇总统计信息（供 UI 展示）
    grand_qty = sum(i['total_qty'] for i in sheet_info_list)
    grand_len = sum(i['total_len'] for i in sheet_info_list)
    stats = {
        'processed_sheets': processed_sheets,
        'grand_qty': grand_qty,
        'grand_len_m': round(grand_len / 1000, 2),   # mm → m
        'spec_stats': {
            spec: {
                'qty': spec_stats[spec][0],
                'len_m': round(spec_stats[spec][1] / 1000, 2),
            }
            for spec in SPEC_ORDER
        },
        'sheet_details': [
            {'title': i['title'], 'qty': i['total_qty'], 'len_m': round(i['total_len'] / 1000, 2)}
            for i in sheet_info_list
        ],
    }
    return output.getvalue(), stats, None


# ══════════════════════════════════════════════════════════════════════════════
# UI 渲染入口（复用 assets/style.css 全局样式，不写任何内联 CSS）
# ══════════════════════════════════════════════════════════════════════════════


def render():
    """Streamlit UI 渲染入口"""
    st.header(":material/functions: 檩条下料单自动合计与规格汇总")
    st.markdown(
        "<p style='color: #64748B; margin-bottom: 20px;'>"
        "自动为每个子表写入三种规格（220*100*3/4/5）的 SUMIF 合计行与总合计行，"
        "并生成跨表汇总统计 Sheet。</p>",
        unsafe_allow_html=True,
    )

    # 步骤指引
    st.markdown("""
        <div class="step-indicator">
            <span>Step 1. 上传下料单</span>
            <span>Step 2. 自动写入合计</span>
            <span>Step 3. 下载结果文件</span>
        </div>
    """, unsafe_allow_html=True)

    # 上传区
    with st.container(border=True):
        st.subheader("数据源上传")
        st.info(
            "提示：Excel 各 Sheet 第一行须包含【数量（根）】【规格（mm）】列标题，"
            "第一列须为可解析为数字的序号列。"
        )
        uploaded_file = st.file_uploader(
            "上传檩条下料单 Excel 文件",
            type=["xlsx", "xls"],
            key="purlins_summary_uploader",
        )

    col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
    with col_btn2:
        process_btn = st.button(
            "开始自动写入合计",
            type="primary",
            use_container_width=True,
        )

    if "purlins_summary_run" not in st.session_state:
        st.session_state.purlins_summary_run = False

    if process_btn:
        st.session_state.purlins_summary_run = True
    if not uploaded_file:
        st.session_state.purlins_summary_run = False
    if not st.session_state.purlins_summary_run:
        return

    try:
        with st.spinner("正在写入合计行并生成汇总统计..."):
            excel_bytes, stats, err = add_total_and_summary(uploaded_file)

        if err:
            st.warning(f"{err}")
            return

        # ── 结果指标卡 ────────────────────────────────────────────────────────
        with st.container(border=True):
            st.subheader("处理结果概览")
            st.markdown(f"""
                <div class="metric-container">
                    <div class="metric-card">
                        <div class="metric-title">处理子表数</div>
                        <div class="metric-value">{stats['processed_sheets']} <span style="font-size:0.9rem;font-weight:normal;color:#64748B;">个</span></div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-title">总根数</div>
                        <div class="metric-value">{int(stats['grand_qty']):,} <span style="font-size:0.9rem;font-weight:normal;color:#64748B;">根</span></div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-title">总长度</div>
                        <div class="metric-value">{stats['grand_len_m']:,.1f} <span style="font-size:0.9rem;font-weight:normal;color:#64748B;">m</span></div>
                    </div>
                </div>
            """, unsafe_allow_html=True)

        # ── 标签页：各子表明细 + 规格统计 + 导出 ─────────────────────────────
        tab_sheets, tab_specs, tab_export = st.tabs([
            "各子表明细", "规格统计", "导出结果"
        ])

        with tab_sheets:
            st.caption(f"共处理 {stats['processed_sheets']} 个子表（已自动忽略无效/缺列 Sheet）")
            rows = [
                {"工作表": d['title'], "总根数": int(d['qty']), "总长度 (m)": d['len_m']}
                for d in stats['sheet_details']
            ]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        with tab_specs:
            st.caption("各规格在全部子表中的汇总统计（内存计算值，供核对，与 Excel 公式结果一致）")
            spec_rows = [
                {
                    "规格 (mm)": spec,
                    "总根数": int(v['qty']),
                    "总长度 (m)": v['len_m'],
                }
                for spec, v in stats['spec_stats'].items()
            ]
            st.dataframe(pd.DataFrame(spec_rows), use_container_width=True, hide_index=True)

        with tab_export:
            st.markdown("### 下载处理结果")
            st.write(
                "输出文件在原有每个子表中写入了三种规格的 **SUMIF 公式合计行**与**总合计行**，"
                "并在首位新增了 **汇总统计** Sheet（含各子表合计与按规格统计两大块）。"
            )
            base_name = uploaded_file.name.rsplit(".", 1)[0]
            st.download_button(
                label="下载合计处理后的 Excel",
                data=excel_bytes,
                file_name=f"{base_name}-合计.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=False,
            )

    except Exception as e:
        st.error(f"处理数据时出错: {e}")
