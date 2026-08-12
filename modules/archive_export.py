# modules/archive_export.py
"""
中建二局标准档案表导出模块
============================
功能：
    1. 将合并后的全量 DataFrame 转置映射为 people.xlsx 规定的标准列格式
    2. 根据工种字段智能识别是否为特殊工种
    3. 拼接家庭住址与紧急联系人/电话为一列
    4. 使用 openpyxl 写入含两行工程标题头的 Excel
    5. 强制将身份证号、银行卡号列设为文本格式 @
"""

from io import BytesIO
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 特殊工种关键字（电工/焊工/架工等需持证上岗工种）
# ─────────────────────────────────────────────────────────────────────────────
SPECIAL_TRADE_KEYWORDS = {
    "电工", "焊工", "架工", "爆破工", "起重工",
    "塔吊", "司机", "信号工", "司索工",
    "高压电", "特种作业", "特殊工种",
    "登高架设", "建筑架子工", "高处作业",
    "气焊", "电焊", "氩弧焊",
    "挖掘机", "装载机", "压路机", "叉车",
}

# 目标档案表标准列名（顺序与 people.xlsx 一致）
ARCHIVE_COLUMNS = [
    "编号",
    "姓名",
    "性别",
    "年龄",
    "民族",
    "籍贯",
    "身份证号",
    "银行卡号",
    "是否为特殊工种",
    "岗位证书编号",
    "手机",
    "家庭住址、家庭联系人及电话",
    "劳务队伍名称",
    "本人职务",
    "工种",
    "入场教育资料存档号",
    "进场时间",
    "退场时间",
    "工资支付情况",
    "不良行为记录",
]

# 强制文本格式 @ 的列名集合
TEXT_FORMAT_COLS = {"身份证号", "银行卡号", "手机"}


def _get(row: pd.Series, *keys: str, default: str = "") -> str:
    """依次尝试多个可能的列名，返回第一个非空值。"""
    for key in keys:
        val = row.get(key, "")
        if pd.notna(val) and str(val).strip() not in ("", "nan", "None", "<NA>"):
            s = str(val).strip()
            # 去除尾部 .0
            if s.endswith(".0"):
                s = s[:-2]
            return s
    return default


def is_special_trade(trade_str: str) -> str:
    """根据工种字符串判断是否为特殊工种，返回 '是' 或 '否'。"""
    if not trade_str or not trade_str.strip():
        return ""
    for kw in SPECIAL_TRADE_KEYWORDS:
        if kw in trade_str:
            return "是"
    return "否"


def _calc_age_from_id(id_no: str) -> str:
    """从18位/15位身份证号推算年龄（无法计算时返回空）。"""
    try:
        import datetime
        id_no = str(id_no).strip()
        if len(id_no) == 18:
            birth_year = int(id_no[6:10])
            birth_month = int(id_no[10:12])
            birth_day = int(id_no[12:14])
        elif len(id_no) == 15:
            birth_year = 1900 + int(id_no[6:8])
            birth_month = int(id_no[8:10])
            birth_day = int(id_no[10:12])
        else:
            return ""
        today = datetime.date.today()
        age = today.year - birth_year - (
            (today.month, today.day) < (birth_month, birth_day)
        )
        return str(age) if 0 < age < 120 else ""
    except Exception:
        return ""


def build_archive_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    将合并后的全量 DataFrame 转置映射为中建二局标准档案格式 DataFrame。

    Parameters
    ----------
    df : pd.DataFrame
        info_merge.process_and_merge() 返回的全量合并表

    Returns
    -------
    pd.DataFrame
        标准列名的档案 DataFrame（列顺序与 ARCHIVE_COLUMNS 一致）
    """
    records = []
    for seq, (_, row) in enumerate(df.iterrows(), start=1):
        id_no = _get(row, "身份证号")
        trade = _get(row, "工种")

        # 12. 住址 + 联系人 + 电话拼接为一列
        address = _get(row, "详细地址", "家庭住址", "住址")
        contact = _get(row, "紧急联系人", "家庭联系人", "联系人")
        contact_tel = _get(row, "紧急联系电话", "联系电话", "家庭电话")

        addr_parts = []
        if address:
            addr_parts.append(address)
        if contact:
            addr_parts.append(f"联系人：{contact}")
        if contact_tel:
            addr_parts.append(f"电话：{contact_tel}")
        full_address_contact = "　".join(addr_parts)  # 全角空格分隔

        # 年龄：优先读字段，缺省则由身份证号推算
        age = _get(row, "年龄")
        if not age and id_no:
            age = _calc_age_from_id(id_no)

        record = {
            "编号": str(seq),
            "姓名": _get(row, "姓名"),
            "性别": _get(row, "性别"),
            "年龄": age,
            "民族": _get(row, "民族"),
            "籍贯": _get(row, "籍贯", "户籍地"),
            "身份证号": id_no,
            "银行卡号": _get(row, "工资卡号", "银行卡号", "卡号"),
            "是否为特殊工种": is_special_trade(trade),
            "岗位证书编号": _get(row, "岗位证书编号", "证书编号", "资质证书编号"),
            "手机": _get(row, "手机号", "手机", "电话", "联系电话"),
            "家庭住址、家庭联系人及电话": full_address_contact,
            "劳务队伍名称": _get(row, "班组", "分包/所属企业", "劳务队伍"),
            "本人职务": _get(row, "人员类型", "职务", "岗位"),
            "工种": trade,
            "入场教育资料存档号": _get(row, "入场教育资料存档号", "入场教育存档号"),
            "进场时间": _get(row, "进场日期", "进场时间", "入场日期"),
            "退场时间": _get(row, "退场日期", "退场时间", "离场日期"),
            "工资支付情况": _get(row, "工资支付情况", "工资发放情况", "发薪情况", "工资结算情况"),
            "不良行为记录": _get(row, "不良行为记录", "违规记录"),
        }
        records.append(record)

    archive_df = pd.DataFrame(records, columns=ARCHIVE_COLUMNS)
    return archive_df


def generate_archive_excel(df: pd.DataFrame, project_name: str = "XX工程劳务人员档案表") -> bytes:
    """
    将合并后的全量 DataFrame 写入中建二局标准档案 Excel，格式说明：
      - 第1行：工程名称标题（跨所有列合并，宋体14号加粗居中）
      - 第2行：列标题头（宋体10号加粗居中，蓝色底色）
      - 第3行起：数据（宋体10号，带边框）
      - 身份证号、银行卡号、手机 列强制设为文本格式 @

    Parameters
    ----------
    df : pd.DataFrame
        info_merge.process_and_merge() 返回的全量合并表（或其筛选子集）
    project_name : str
        工程名称，写入第1行标题，默认为"XX工程劳务人员档案表"

    Returns
    -------
    bytes
        Excel 文件二进制内容
    """
    archive_df = build_archive_df(df)
    headers = list(archive_df.columns)
    col_count = len(headers)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "中建二局劳务档案"

    # ── 样式定义 ──────────────────────────────────────────────────────────────
    title_font = Font(name="宋体", size=14, bold=True)
    header_font = Font(name="宋体", size=10, bold=True)
    data_font = Font(name="宋体", size=10)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    header_fill = PatternFill(fill_type="solid", fgColor="DAEEF3")

    thin_side = Side(style="thin")
    thin_border = Border(
        left=thin_side, right=thin_side,
        top=thin_side, bottom=thin_side
    )

    # ── 第1行：工程标题 ────────────────────────────────────────────────────────
    ws.append([project_name] + [""] * (col_count - 1))
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = title_font
    title_cell.alignment = center_align
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=col_count
    )
    ws.row_dimensions[1].height = 36

    # ── 第2行：列标题 ─────────────────────────────────────────────────────────
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
    ws.row_dimensions[2].height = 24

    # ── 计算需文本格式的列索引（1-based）────────────────────────────────────
    text_col_indices = {
        idx + 1
        for idx, col in enumerate(headers)
        if col in TEXT_FORMAT_COLS
    }

    # ── 第3行起：数据 ─────────────────────────────────────────────────────────
    for data_row in archive_df.itertuples(index=False):
        row_vals = list(data_row)
        ws.append(row_vals)
        current_row = ws.max_row
        ws.row_dimensions[current_row].height = 18

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            col_name = headers[col_idx - 1]

            # 长文本列左对齐，其余居中
            if col_name == "家庭住址、家庭联系人及电话":
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            # 强制文本格式，防止科学计数法
            if col_idx in text_col_indices:
                cell.number_format = "@"
                cell.value = str(val) if val is not None else ""

    # ── 列宽预设 ──────────────────────────────────────────────────────────────
    col_width_map = {
        "编号": 6,
        "姓名": 10,
        "性别": 6,
        "年龄": 6,
        "民族": 8,
        "籍贯": 14,
        "身份证号": 20,
        "银行卡号": 22,
        "是否为特殊工种": 14,
        "岗位证书编号": 20,
        "手机": 14,
        "家庭住址、家庭联系人及电话": 42,
        "劳务队伍名称": 18,
        "本人职务": 12,
        "工种": 12,
        "入场教育资料存档号": 22,
        "进场时间": 13,
        "退场时间": 13,
        "工资支付情况": 16,
        "不良行为记录": 18,
    }
    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_width_map.get(col_name, 14)

    # ── 冻结表头（第2行以下可滚动）───────────────────────────────────────────
    ws.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_archive_excel_multi_sheet(
    df: pd.DataFrame,
    project_name: str = "XX工程劳务人员档案表",
    group_col: str = "分包/所属企业",
) -> bytes:
    """
    按 group_col（默认：分包/所属企业）将数据分组，每个公司写入独立的 Sheet，
    格式与 generate_archive_excel 完全一致（标题行、列头、边框、列宽、冻结行）。

    Parameters
    ----------
    df          : 合并后的全量 DataFrame
    project_name: 工程名称，写入每个 Sheet 的第1行标题
    group_col   : 用于分组的列名，默认 '分包/所属企业'
    """
    FALLBACK = "未分配企业"

    # ── 样式定义（与 generate_archive_excel 保持一致）──────────────────────
    title_font   = Font(name="宋体", size=14, bold=True)
    header_font  = Font(name="宋体", size=10, bold=True)
    data_font    = Font(name="宋体", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    header_fill  = PatternFill(fill_type="solid", fgColor="DAEEF3")
    thin_side    = Side(style="thin")
    thin_border  = Border(left=thin_side, right=thin_side,
                          top=thin_side, bottom=thin_side)

    col_width_map = {
        "编号": 6, "姓名": 10, "性别": 6, "年龄": 6, "民族": 8, "籍贯": 14,
        "身份证号": 20, "银行卡号": 22, "是否为特殊工种": 14, "岗位证书编号": 20,
        "手机": 14, "家庭住址、家庭联系人及电话": 42, "劳务队伍名称": 18,
        "本人职务": 12, "工种": 12, "入场教育资料存档号": 22,
        "进场时间": 13, "退场时间": 13, "工资支付情况": 16, "不良行为记录": 18,
    }

    # ── 确定分组 ──────────────────────────────────────────────────────────
    if group_col in df.columns:
        groups = df[group_col].fillna(FALLBACK).replace("", FALLBACK).unique()
    else:
        groups = [FALLBACK]

    wb = openpyxl.Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    def _write_archive_sheet(ws, company_df, sheet_project_name):
        archive_df = build_archive_df(company_df)
        headers    = list(archive_df.columns)
        col_count  = len(headers)

        text_col_indices = {
            idx + 1 for idx, col in enumerate(headers) if col in TEXT_FORMAT_COLS
        }

        # 第1行：工程标题
        ws.append([sheet_project_name] + [""] * (col_count - 1))
        cell = ws.cell(row=1, column=1)
        cell.font      = title_font
        cell.alignment = center_align
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws.row_dimensions[1].height = 36

        # 第2行：列标题
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            c = ws.cell(row=2, column=col_idx)
            c.font      = header_font
            c.alignment = center_align
            c.fill      = header_fill
            c.border    = thin_border
        ws.row_dimensions[2].height = 24

        # 第3行起：数据
        for data_row in archive_df.itertuples(index=False):
            row_vals = list(data_row)
            ws.append(row_vals)
            cur_row = ws.max_row
            ws.row_dimensions[cur_row].height = 18
            for col_idx, val in enumerate(row_vals, start=1):
                c          = ws.cell(row=cur_row, column=col_idx)
                c.font     = data_font
                c.border   = thin_border
                col_name   = headers[col_idx - 1]
                c.alignment = left_align if col_name == "家庭住址、家庭联系人及电话" else center_align
                if col_idx in text_col_indices:
                    c.number_format = "@"
                    c.value = str(val) if val is not None else ""

        # 列宽
        for col_idx, col_name in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width_map.get(col_name, 14)

        # 冻结表头
        ws.freeze_panes = "A3"

    for company in groups:
        if group_col in df.columns:
            mask = (df[group_col] == company) | \
                   (df[group_col].fillna(FALLBACK).replace("", FALLBACK) == company)
            company_df = df[mask]
        else:
            company_df = df

        if company_df.empty:
            continue

        # Sheet 名称（最多31字符，避免重名）
        sheet_name = str(company).strip()[:31] or FALLBACK
        base_name  = sheet_name
        counter    = 1
        while sheet_name in wb.sheetnames:
            suffix     = f"_{counter}"
            sheet_name = f"{base_name[:31-len(suffix)]}{suffix}"
            counter   += 1

        ws = wb.create_sheet(title=sheet_name)
        _write_archive_sheet(ws, company_df, project_name)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
