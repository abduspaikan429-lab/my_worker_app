# modules/info_merge.py
import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
import os
from modules import archive_export
from modules.master_data import load_master_df, preview_update, commit_update, get_last_changes

def clean_val(val):
    """防止浮点数带.0以及处理科学计数法字符串与多余空格"""
    if pd.isna(val) or val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ["nan", "none", "null", "<na>"]:
        return ""
    if s.endswith('.0'):
        s = s[:-2]
    if 'e+' in s.lower() or 'e-' in s.lower():
        try:
            s = f"{float(s):.0f}"
        except Exception:
            pass
    return s

def smart_read_excel(file, dtype=str):
    """
    智能读取 Excel 文件：自动跳过前几行非表格内容（如项目名称、导出时间等），
    定位含有「身份证号」或「姓名」的行作为真正表头，再截取后续有效数据。
    """
    HEADER_KEYWORDS = ['身份证号', '姓名']
    MAX_SCAN_ROWS = 30  # 最多向下扫描的行数，防止恶意大文件

    # 确保文件指针从头开始，防止多次调用时读到空数据
    if hasattr(file, 'seek'):
        file.seek(0)
    raw = pd.read_excel(file, header=None, dtype=dtype)
    header_row_idx = None
    for idx in range(min(MAX_SCAN_ROWS, len(raw))):
        row_vals = [str(v).strip() for v in raw.iloc[idx].tolist()]
        if any(kw in row_vals for kw in HEADER_KEYWORDS):
            header_row_idx = idx
            break

    if header_row_idx is None:
        if hasattr(file, 'seek'):
            file.seek(0)
        return pd.read_excel(file, dtype=dtype)

    new_columns = [str(v).replace('\n', '').strip() for v in raw.iloc[header_row_idx].tolist()]
    data_rows = raw.iloc[header_row_idx + 1:].reset_index(drop=True)
    data_rows.columns = new_columns
    data_rows = data_rows.dropna(how='all').reset_index(drop=True)
    return data_rows


def process_and_merge(files_a, files_b):
    """
    接收两个 Excel 文件流列表，执行数据清洗、同义映射、外连接合并与去重
    """
    def process_file_list(file_list, sys_type):
        all_dfs = []
        if not file_list:
            return None
        
        mapping = {
            '劳动合同': '合同签订状态',
            '合同签订': '合同签订状态',
            '工资标准': '结算单价/标准',
            '结算单价(元)': '结算单价/标准',
            '银行': '开户银行',
            '工资卡银行': '开户银行',
            '家庭住址': '详细地址',
            '分包单位': '分包/所属企业',
            '所属企业': '分包/所属企业',
            '进退场状态': '在场/进退场状态',
            '在场情况': '在场/进退场状态',
            '工人类型': '人员类型',
            '所属班组': '班组',
        }
        
        for file in file_list:
            df = smart_read_excel(file, dtype=str)
            if df is not None and not df.empty:
                df.columns = [str(c).replace('\n', '').strip() for c in df.columns]
                if sys_type == 'B' and '序号' in df.columns:
                    df = df.drop(columns=['序号'])
                df = df.rename(columns=mapping)
                if '身份证号' in df.columns:
                    df['身份证号'] = df['身份证号'].apply(clean_val)
                    df = df[df['身份证号'].str.len() >= 15]
                all_dfs.append(df)
        
        if not all_dfs:
            return None
        
        combined_df = pd.concat(all_dfs, ignore_index=True)
        if '身份证号' in combined_df.columns:
            combined_df = combined_df.drop_duplicates(subset=['身份证号'], keep='last')
        return combined_df

    df_a = process_file_list(files_a, 'A')
    df_b = process_file_list(files_b, 'B')

    if df_a is None and df_b is None:
        return None, 0, 0, 0

    if df_a is None:
        for c in df_b.columns:
            df_b[c] = df_b[c].apply(clean_val)
        return df_b, len(df_b), 0, len(df_b)

    if df_b is None:
        for c in df_a.columns:
            df_a[c] = df_a[c].apply(clean_val)
        return df_a, len(df_a), 0, len(df_a)

    # 统计重合人数
    set_a = set(df_a['身份证号'].dropna().unique()) if '身份证号' in df_a.columns else set()
    set_b = set(df_b['身份证号'].dropna().unique()) if '身份证号' in df_b.columns else set()
    overlap_count = len(set_a.intersection(set_b))

    # Outer Join 外连接
    merged = pd.merge(df_a, df_b, on='身份证号', how='outer', suffixes=('_A', '_B'))

    # 保留全量字段
    all_cols = []
    for col in list(df_a.columns) + list(df_b.columns):
        if col != '身份证号' and col not in all_cols:
            all_cols.append(col)

    final_df = pd.DataFrame()
    final_df['身份证号'] = merged['身份证号']

    for col in all_cols:
        col_a = f"{col}_A"
        col_b = f"{col}_B"
        if col_a in merged.columns and col_b in merged.columns:
            final_df[col] = merged[col_a].replace('', pd.NA).fillna(merged[col_b])
        elif col_a in merged.columns:
            final_df[col] = merged[col_a]
        elif col_b in merged.columns:
            final_df[col] = merged[col_b]
        elif col in merged.columns:
            final_df[col] = merged[col]

    # 清理所有字段格式
    for c in final_df.columns:
        final_df[c] = final_df[c].apply(clean_val)

    # 智能列重排
    desired_order = [
        # 1. 个人基础信息
        "姓名", "性别", "民族", "年龄", "身份证号", "手机号", "详细地址", "家庭住址",
        # 2. 进场与班组信息
        "班组", "工种", "人员类型", "进场日期", "进场时间", "在场状态", "进退场状态", "在场/进退场状态",
        # 3. 银行与发薪信息
        "银行卡号", "工资卡号", "开户银行",
        # 4. 合同与合规信息
        "劳动合同编号", "合同签订状态", "劳动合同", "是否在市建委"
    ]
    
    existing_cols = [c for c in desired_order if c in final_df.columns]
    remaining_cols = [c for c in final_df.columns if c not in existing_cols]
    ordered_df = final_df[existing_cols + remaining_cols]

    return ordered_df, len(set_a), overlap_count, len(set_b)

def generate_excel(df):
    """
    使用 openpyxl 按班组分 Sheet 导出，绘制中建标准复杂表头，且设置纯文本格式 @ 单元格
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Extract unique companies (分包/所属企业) for sheet grouping
    GROUP_COL = '分包/所属企业'
    FALLBACK_NAME = "未分配企业"
    if GROUP_COL in df.columns:
        teams = df[GROUP_COL].fillna(FALLBACK_NAME).replace("", FALLBACK_NAME).unique()
    else:
        teams = [FALLBACK_NAME]
        
    if len(teams) == 0:
        teams = [FALLBACK_NAME]

    text_keywords = ['身份证', '手机', '电话', '卡号', '银行卡', '工资卡', '编号', '代码']
    headers = list(df.columns)
    # 序号列占第1列，原字段列从第2列开始，索引+2
    text_col_indices = [
        idx + 2 for idx, col in enumerate(headers)
        if any(kw in col for kw in text_keywords)
    ]
    display_headers = ["序号"] + headers
    
    # Styles
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for team in teams:
        sheet_name = str(team).strip()[:31]
        if not sheet_name:
            sheet_name = "未分配班组"
            
        # Ensure unique sheet names if truncation causes duplicates
        base_sheet_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            suffix = f"_{counter}"
            sheet_name = f"{base_sheet_name[:31-len(suffix)]}{suffix}"
            counter += 1
            
        ws = wb.create_sheet(title=sheet_name)
        
        if GROUP_COL in df.columns:
            team_df = df[
                (df[GROUP_COL] == team) |
                (df[GROUP_COL].fillna(FALLBACK_NAME).replace("", FALLBACK_NAME) == team)
            ]
        else:
            team_df = df

        # Fix 12: 跳过空班组，避免生成无数据的空 Sheet
        if team_df.empty:
            wb.remove(ws)
            continue

        max_col = len(display_headers)
        max_col_letter = get_column_letter(max_col) if max_col > 0 else 'A'
        
        # Row 1, 2 heights and Col A width
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30
        ws.column_dimensions['A'].width = 15
        
        # Logo — 使用绝对路径，避免因工作目录不同而找不到文件
        ws.merge_cells('A1:A2')
        _base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        logo_path = os.path.join(_base_dir, 'assets', 'cscec_logo.png')
        if os.path.exists(logo_path):
            try:
                img = OpenpyxlImage(logo_path)
                img.width = 80
                img.height = 70
                ws.add_image(img, 'A1')
            except Exception:
                pass
                
        # Main Title
        if max_col >= 2:
            main_title_range = f'B1:{max_col_letter}1'
            ws.merge_cells(main_title_range)
            ws['B1'] = '中国建筑  管理表格'
            ws['B1'].font = Font(name='黑体', size=18, bold=True)
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            
        # Sub Title and Table No
        if max_col >= 4:
            sub_title_range = f'B2:{get_column_letter(max_col-2)}2'
            ws.merge_cells(sub_title_range)
            ws['B2'] = '人员信息档案表'
            ws['B2'].font = Font(name='黑体', size=16, bold=True)
            ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
            
            table_no_range = f'{get_column_letter(max_col-1)}2:{max_col_letter}2'
            ws.merge_cells(table_no_range)
            ws[f'{get_column_letter(max_col-1)}2'] = '表格编号                  '
            ws[f'{get_column_letter(max_col-1)}2'].font = Font(size=11)
            ws[f'{get_column_letter(max_col-1)}2'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Apply borders to title
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = thin_border
        
        # Headers (Row 3)
        ws.append(display_headers)
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        # Data — 每个 Sheet 序号从 1 重新开始
        for seq_num, row in enumerate(team_df.itertuples(index=False), start=1):
            ws.append([seq_num] + [clean_val(v) for v in row])
            
        # Format text columns and apply borders to data
        start_row = 4
        end_row = len(team_df) + 3
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
            for col_idx in text_col_indices:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = '@'
                if cell.value is not None:
                    cell.value = str(cell.value)
                    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def render():
    """简洁现代化主界面"""
    st.markdown("""
    <div class="page-header-deco">
        <span class="header-emoji">📋</span>
        <div class="header-text">
            <h2>劳务人员多系统档案自动整合</h2>
            <p>将多平台导出的劳务档案标准化清洗、去重与合并输出</p>
        </div>
    </div>
    <div class="color-strip"></div>
    """, unsafe_allow_html=True)

    # 1. 交互流：清晰的步骤指引
    st.markdown("""
        <div class="step-indicator">
            <span class="step-num">1</span>
            <span>上传源数据文件</span>
            <span class="step-num">2</span>
            <span>自动清洗与合并</span>
            <span class="step-num">3</span>
            <span>预览与导出</span>
        </div>
    """, unsafe_allow_html=True)

    # 2. 上传区域卡片化 (支持多文件上传)
    with st.container(border=True):
        st.markdown("#### 数据源上传")
        st.markdown('<div class="hint-box">提示：支持多选文件批量上传。请确保上传的 Excel 表格包含『身份证号』列，系统将以此为基准进行档案匹配和去重。</div>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            files_a = st.file_uploader("三局智慧工地系统导出表 (多选)", type=["xlsx", "xls"], key="files_a", accept_multiple_files=True)
        with col2:
            files_b = st.file_uploader("智慧护薪系统导出表 (多选)", type=["xlsx", "xls"], key="files_b", accept_multiple_files=True)

    # 3. 主操作按钮
    col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
    with col_btn2:
        process_btn = st.button("开始智能清洗与合并", type="primary", use_container_width=True)

    if 'info_merge_run' not in st.session_state:
        st.session_state.info_merge_run = False
        
    if process_btn:
        st.session_state.info_merge_run = True
        
    if not files_a and not files_b and 'merged_df' not in st.session_state:
        st.session_state.info_merge_run = False

    if not st.session_state.info_merge_run and 'merged_df' not in st.session_state:
        return

    try:
        if process_btn:
            with st.spinner("数据清洗中..."):
                result_df, count_a, overlap_count, count_b = process_and_merge(files_a, files_b)

            if result_df is None or result_df.empty:
                st.warning("未能提取到有效的劳务人员数据。")
                return
            st.session_state.merged_df = result_df
            st.session_state.info_merge_counts = (count_a, overlap_count, count_b)
            st.session_state.master_sync_result = None

        result_df = st.session_state.get('merged_df')
        if result_df is None or result_df.empty:
            st.warning("未能提取到有效的劳务人员数据。")
            return
        count_a, overlap_count, count_b = st.session_state.get('info_merge_counts', (0, 0, 0))

        # 官网导出结果先进入预览，确认后才写入项目人员主表。
        master_preview = preview_update(load_master_df(), result_df)
        if master_preview.get('error'):
            st.error(master_preview['error'])
        else:
            new_count = len(master_preview['new_rows'])
            updated_count = len(master_preview['updated_rows'])
            missing_count = len(master_preview['missing_from_import'])
            st.markdown("### 官网数据同步到项目主表")
            st.markdown(
                '<div class="hint-box">进场流程只负责新人员手续跟踪。只有智慧护薪和三局系统正式建档后，才在这里同步人员主表。空白官网字段不会覆盖主表已有资料。</div>',
                unsafe_allow_html=True,
            )
            sm1, sm2, sm3, sm4 = st.columns(4)
            sm1.metric("本次新增", f"{new_count} 人")
            sm2.metric("资料变化", f"{updated_count} 人")
            sm3.metric("主表同步后", f"{len(master_preview['merged_df'])} 人")
            sm4.metric("本次未出现", f"{missing_count} 人")
            if missing_count:
                st.caption("本次官网导出中未出现的人员不会被删除，系统只提示，不做破坏性处理。")

            if new_count or updated_count:
                with st.expander("预览本次会写入主表的人员变化", expanded=True):
                    if new_count:
                        st.markdown("**新增人员**")
                        st.dataframe(master_preview['new_rows'], use_container_width=True, hide_index=True)
                    if updated_count:
                        st.markdown("**信息变化人员**")
                        st.dataframe(master_preview['updated_rows'], use_container_width=True, hide_index=True)
                if st.button("确认同步，并生成新增人员直贴数据", type="primary", use_container_width=True, key="confirm_master_sync"):
                    saved = commit_update(
                        result_df,
                        source_files=[getattr(f, 'name', '') for f in (files_a or []) + (files_b or [])],
                    )
                    if saved.get('error'):
                        st.error(saved['error'])
                    else:
                        st.session_state.merged_df = saved['merged_df']
                        st.session_state.master_sync_result = saved
                        st.success(f"主表已同步。新增 {len(saved['new_rows'])} 人，资料变化 {len(saved['updated_rows'])} 人。请到【花名册与报表】直接复制新增行。")
            elif st.session_state.get('master_sync_result'):
                saved = st.session_state.master_sync_result
                st.success(f"最近一次同步完成：新增 {len(saved.get('new_rows', []))} 人，资料变化 {len(saved.get('updated_rows', []))} 人。")

        # 计算重合率
        total_unique = len(result_df)
        overlap_rate = f"{(overlap_count / total_unique * 100):.1f}%" if total_unique > 0 else "0%"

        # 4. 核心指标卡片
        st.markdown("""
        <div class="celebrate-banner">
            <span>数据处理完成！以下是合并结果概览</span>
        </div>
        """, unsafe_allow_html=True)

        with st.container(border=True):
            st.markdown("#### 处理结果概览")
            st.markdown(f"""
                <div class="metric-container">
                    <div class="metric-card">
                        <div class="metric-title">汇总人数</div>
                        <div class="metric-value">{total_unique} <span style="font-size:0.9rem; font-weight:normal; color:#64748B;">人</span></div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-title">双系统重合人数</div>
                        <div class="metric-value">{overlap_count} <span style="font-size:0.9rem; font-weight:normal; color:#64748B;">人</span></div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-title">双系统重合率</div>
                        <div class="metric-value" style="color: #F59E0B;">{overlap_rate}</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-title">全量保留字段</div>
                        <div class="metric-value">{len(result_df.columns)} <span style="font-size:0.9rem; font-weight:normal; color:#64748B;">列</span></div>
                    </div>
                </div>
            """, unsafe_allow_html=True)

        # 5. 分标签页展示：检索、统计分析、导出
        tab_data, tab_charts, tab_export = st.tabs(["人员档案明细与检索", "工种与班组统计", "数据导出"])

        # 初始化检索 Session State
        if 'search_query' not in st.session_state:
            st.session_state.search_query = ""
        if 'team_filter' not in st.session_state:
            st.session_state.team_filter = "全部班组"
        if 'trade_filter' not in st.session_state:
            st.session_state.trade_filter = "全部工种"

        def reset_filters():
            st.session_state.search_query = ""
            st.session_state.team_filter = "全部班组"
            st.session_state.trade_filter = "全部工种"

        with tab_data:
            # 筛选控制工具栏
            f1, f2, f3, f4 = st.columns([2.5, 1.2, 1.2, 0.8])

            with f1:
                search_key = st.text_input("搜索", value=st.session_state.search_query, placeholder="输入姓名或身份证号搜索...", label_visibility="collapsed")
                st.session_state.search_query = search_key

            with f2:
                team_options = ["全部班组"]
                if '班组' in result_df.columns:
                    team_options.extend(sorted([t for t in result_df['班组'].unique() if t]))
                selected_team = st.selectbox("班组", team_options, index=team_options.index(st.session_state.team_filter) if st.session_state.team_filter in team_options else 0, label_visibility="collapsed")
                st.session_state.team_filter = selected_team

            with f3:
                trade_options = ["全部工种"]
                if '工种' in result_df.columns:
                    trade_options.extend(sorted([t for t in result_df['工种'].unique() if t]))
                selected_trade = st.selectbox("工种", trade_options, index=trade_options.index(st.session_state.trade_filter) if st.session_state.trade_filter in trade_options else 0, label_visibility="collapsed")
                st.session_state.trade_filter = selected_trade

            with f4:
                st.button("重置", on_click=reset_filters, use_container_width=True)

            # 数据过滤
            filtered_df = result_df

            if search_key:
                name_mask = filtered_df['姓名'].astype(str).str.contains(search_key, na=False) if '姓名' in filtered_df.columns else False
                id_mask = filtered_df['身份证号'].astype(str).str.contains(search_key, na=False) if '身份证号' in filtered_df.columns else False
                filtered_df = filtered_df[name_mask | id_mask]

            if selected_team != "全部班组" and '班组' in filtered_df.columns:
                filtered_df = filtered_df[filtered_df['班组'] == selected_team]

            if selected_trade != "全部工种" and '工种' in filtered_df.columns:
                filtered_df = filtered_df[filtered_df['工种'] == selected_trade]

            # 显示结果统计
            st.caption(f"显示 **{len(filtered_df)}** / {len(result_df)} 条记录")

            # 展出表格
            st.dataframe(filtered_df, use_container_width=True, height=600)

        with tab_charts:
            c1, c2 = st.columns(2)
            with c1:
                st.subheader("工种分布分析")
                if '工种' in filtered_df.columns and not filtered_df['工种'].replace('', pd.NA).dropna().empty:
                    trade_counts = filtered_df['工种'].value_counts()
                    st.bar_chart(trade_counts)
                else:
                    st.info("暂无工种数据")

            with c2:
                st.subheader("班组分布分析")
                if '班组' in filtered_df.columns and not filtered_df['班组'].replace('', pd.NA).dropna().empty:
                    team_counts = filtered_df['班组'].value_counts()
                    st.bar_chart(team_counts)
                else:
                    st.info("暂无班组数据")

        with tab_export:
            st.markdown("### 导出整合后的 Excel 文件")

            # ── 动态计算企业名称和文件名（按分包/所属企业）──────────────────
            GROUP_COL = '分包/所属企业'
            teams = []
            if GROUP_COL in filtered_df.columns:
                teams = sorted([t for t in filtered_df[GROUP_COL].fillna("未分配").replace("", "未分配").unique() if str(t).strip() not in ("", "未分配")])
            
            if len(teams) == 0:
                team_suffix = "全量"
            elif len(teams) <= 3:
                team_suffix = "_".join([str(t)[:10] for t in teams])
            else:
                team_suffix = f"{len(teams)}个班组"

            # ── 导出模式选择 ──────────────────────────────────────────────
            export_mode = st.radio(
                "选择导出格式",
                options=["📊 系统导出人员信息整理汇总", "📋 人员信息档案表（中建二局标准）"],
                horizontal=True,
                key="export_mode_radio",
            )

            if export_mode == "📊 系统导出人员信息整理汇总":
                # ── 全量信息表：按班组分 Sheet，含中建标准复杂表头 ──────────
                st.caption("包含所有原始字段的完整档案表，按班组自动分为多个 Sheet，带标准跨列标题头与企业 Logo，适用于内部存档与数据核查。")
                st.markdown('<div class="hint-box">已使用中建标准模板格式，按班组自动分为多个 Sheet，带有标准的跨列标题头与企业 Logo。同时身份证号、手机号、工资卡号等关键数据已被强制格式化为纯文本(@)，防止出现科学计数法。</div>', unsafe_allow_html=True)

                excel_bytes = generate_excel(filtered_df)
                file_name = f"劳务人员汇总_{team_suffix}.xlsx"

                st.download_button(
                    label="📥 导出全量人员信息汇总表 Excel（多 Sheet 按班组）",
                    data=excel_bytes,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=False,
                )

            else:
                # ── 中建二局标准档案表 ────────────────────────────────────
                st.caption(
                    "按中建二局 people.xlsx 标准格式输出，含工程标题头、特殊工种自动识别、"
                    "住址与联系人拼接，身份证号/银行卡号强制文本格式 @。"
                )
                project_name = st.text_input(
                    "工程名称（将写入表格第一行标题）",
                    value="XX工程劳务人员档案表",
                    key="archive_project_name",
                    placeholder="请输入工程全称，例如：XX项目劳务人员档案表",
                )

                # 实时预览字段映射结果
                with st.expander("预览档案表格式（前5行）", expanded=False):
                    try:
                        preview_df = archive_export.build_archive_df(filtered_df)
                        st.dataframe(preview_df.head(5), use_container_width=True)
                    except Exception as prev_err:
                        st.warning(f"预览失败：{prev_err}")

                archive_bytes = archive_export.generate_archive_excel_multi_sheet(
                    filtered_df,
                    project_name=project_name or "XX工程劳务人员档案表",
                    group_col="分包/所属企业",
                )
                archive_file_name = f"{project_name or '劳务人员档案表'}_{team_suffix}.xlsx"

                st.download_button(
                    label="📥 导出人员信息档案表 Excel（中建二局标准）",
                    data=archive_bytes,
                    file_name=archive_file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=False,
                )

    except Exception as e:
        st.error(f"处理数据时出错: {e}")
