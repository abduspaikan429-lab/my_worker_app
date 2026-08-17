# modules/attendance_payroll.py
# ============================================================
# 考勤对账、定稿与工资联动结算模块  v2
# 架构: 三层解耦 — 数据层 / 导出层 / 展示层
# CSS: 严格复用 assets/style.css，禁止内联硬编码样式
#
# Tab 1 流程门控设计:
#   Step A -> 三方差异对比报告（只读，供领导确认）
#   Step B -> 定稿录入（领导确认后，录入最终核定天数）
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import os
from copy import copy
from pathlib import Path
from datetime import datetime
from modules.master_data import load_master_df

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ================================================================
# 数据层
# ================================================================

# ── 列名候选表（精确匹配单元格值）─────────────────────────────
_NAME_CANDS = ['姓名', '名', '名字', '人员姓名', 'Name']
_ID_CANDS = ['身份证号', '身份证号码', '身份证', '证件号码', '证件号']
_COMPANY_CANDS = ['分包/所属企业', '所属企业', '分包单位', '单位', '劳务单位']
_DAYS_CANDS = [
    '考勤天数', '出勤天数', '实际天数', '天数', '实出勤', '实际出勤天数',
    '当月出勤天数', '小计', '合计', '合计天数', '总计', '天', '出勤合计', '出勤', 'Days',
]
_SERIAL_CANDS = ['序号', '编号', '序', 'No.', 'NO.']
_NAME_EXCLUDED_HINTS = (
    '签字', '签名', '申明', '声明', '项目', '工程', '班组', '负责人',
    '名称', '编号', '序号', '日期', '备注', '单位', '工资', '考勤', '人员',
)


def _normalize_header(value):
    """统一处理 Excel 表头中的换行、普通空格和全角空格。"""
    return re.sub(r'[\s\u3000]+', '', str(value).strip())


def _clean_identity(value):
    """清理身份证号中的空格和 Excel 文本化产生的尾部 .0。"""
    text = _raw_attendance_value(value).replace(' ', '').upper()
    return text[:-2] if text.endswith('.0') else text


def _is_valid_identity(value):
    return len(_clean_identity(value)) >= 15


def _looks_like_person_name(value):
    """过滤模板里的声明、签字、项目名称等非人员文本。"""
    name = _raw_attendance_value(value)
    compact = _normalize_header(name)
    if not compact or any(hint in compact for hint in _NAME_EXCLUDED_HINTS):
        return False
    return bool(re.fullmatch(r'[\u4e00-\u9fff·]{2,12}', compact))


def _detect_header_row(file_obj) -> int:
    """
    扫描前 12 行，找到包含姓名字段的真实表头行。
    兼容“日期\n姓名”“姓名（签字）”等组合表头。
    """
    try:
        raw = pd.read_excel(file_obj, dtype=str, header=None, nrows=12)
        for i in range(min(12, len(raw))):
            row_cells = [_normalize_header(v) for v in raw.iloc[i].tolist()]
            if any(
                ('姓名' in cell or '人员姓名' in cell or '名字' in cell or cell == '名' or cell.lower() == 'name')
                for cell in row_cells
            ):
                return i
    except Exception:
        pass
    return 0


def _safe_read(file_obj, label: str) -> pd.DataFrame:
    """
    智能读取考勤 Excel：
    1. 自动检测真实表头行（精确匹配姓名候选词）
    2. 统一字符串类型、清洗列名
    3. 过滤空行和纯数字序号/分隔行
    """
    if file_obj is None:
        return pd.DataFrame()
    try:
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
        header_row = _detect_header_row(file_obj)

        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)
        df = pd.read_excel(file_obj, dtype=str, header=header_row)
        df.columns = [re.sub(r'[\s\u3000]+', '', str(c).strip()) for c in df.columns]

        name_col = _resolve_col(df, _NAME_CANDS)
        if name_col:
            id_col = _resolve_col(df, _ID_CANDS)
            serial_col = _resolve_col(df, _SERIAL_CANDS)

            def _is_person_row(row):
                name = row.get(name_col, '')
                if _is_valid_identity(row.get(id_col, '')) if id_col else False:
                    return True
                serial = str(row.get(serial_col, '')).strip() if serial_col else ''
                return _looks_like_person_name(name) and (bool(re.fullmatch(r'\d+(?:\.0)?', serial)) or not serial_col)

            mask = (
                df[name_col].notna() &
                (df[name_col].astype(str).str.strip() != '') &
                df.apply(_is_person_row, axis=1)
            )
            df = df[mask].reset_index(drop=True)

        df['__来源__'] = label
        id_col = _resolve_col(df, _ID_CANDS)
        if id_col and id_col != '身份证号':
            df['身份证号'] = df[id_col].apply(_clean_identity)
        elif id_col:
            df['身份证号'] = df[id_col].apply(_clean_identity)
        company_col = _resolve_col(df, _COMPANY_CANDS)
        if company_col and company_col != '分包/所属企业':
            df['分包/所属企业'] = df[company_col].astype(str).str.strip()
        return df
    except Exception as exc:
        st.error(f'读取【{label}】文件失败：{exc}')
        return pd.DataFrame()


def _resolve_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    normalized_columns = {
        column: _normalize_header(column)
        for column in df.columns
    }
    # 兼容组合表头，例如“日期姓名”“姓名（签字）”；单字“名”只允许精确匹配，
    # 防止把“项目名称”“班组名称”误认为姓名列。
    for candidate in candidates:
        normalized_candidate = _normalize_header(candidate)
        for column, normalized_column in normalized_columns.items():
            if normalized_candidate == '名':
                if normalized_column == '名':
                    return column
            elif normalized_candidate and normalized_candidate in normalized_column:
                return column
    return None


def _day_number(column):
    """识别考勤表中的日期列：1、01、1日、01日等均视为同一天。"""
    text = str(column).strip().replace("号", "").replace("日", "")
    if text.isdigit():
        day = int(text)
        if 1 <= day <= 31:
            return day
    return None


def _raw_attendance_value(value):
    """保留官网/纸质表原始内容，不预设最终输出为√或时间段。"""
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null", "<na>"}:
        return ""
    return text


def _attendance_state(value):
    """仅做比对状态识别，不决定最终写回工资表的格式。"""
    text = _raw_attendance_value(value)
    if not text or text in {"--", "—", "-", "/", "0"}:
        return "缺勤/无记录"
    if any(mark in text for mark in ["√", "✓", "✔", "出勤", "正常"]):
        return "有考勤"
    if re.search(r"\d{1,2}:\d{2}", text):
        return "有考勤"
    try:
        if float(text) > 0:
            return "有考勤"
    except (ValueError, TypeError):
        pass
    return "待确认"


def _extract_daily_source(df, source_name):
    if df is None or df.empty:
        return {}
    name_col = _resolve_col(df, _NAME_CANDS)
    id_col = _resolve_col(df, _ID_CANDS)
    if name_col is None:
        return {}

    day_cols = {day: col for col in df.columns if (day := _day_number(col)) is not None}
    records = {}
    for _, row in df.iterrows():
        name = _raw_attendance_value(row.get(name_col, ""))
        identity = _raw_attendance_value(row.get(id_col, "")) if id_col else ""
        key = identity if len(identity) >= 15 else f"姓名:{name}"
        if not name or not key:
            continue
        item = records.setdefault(
            key,
            {"姓名": name, "身份证号": identity, "来源": source_name, "days": {}},
        )
        if not item["身份证号"] and identity:
            item["身份证号"] = identity
        for day, col in day_cols.items():
            value = _raw_attendance_value(row.get(col, ""))
            if value:
                item["days"][day] = value
    return records


def build_daily_diff_table(df_a, df_b, df_paper, company=""):
    """按人员+日期保留三份考勤原始单元格，供人工确认差异。

    这里故意不生成“最终√”或“最终时间段”，只保存原始值和识别状态，
    等领导确认后再决定最终输出格式。
    """
    sources = [
        ("三局", _extract_daily_source(df_a, "三局")),
        ("护薪", _extract_daily_source(df_b, "护薪")),
        ("纸质", _extract_daily_source(df_paper, "纸质")),
    ]
    all_keys = set()
    for _, records in sources:
        all_keys.update(records.keys())

    rows = []
    for key in sorted(all_keys):
        records = {name: data.get(key, {}) for name, data in sources}
        base = next((data for data in records.values() if data), {})
        for day in range(1, 32):
            values = {name: data.get("days", {}).get(day, "") for name, data in records.items()}
            states = [_attendance_state(values[name]) for name in ["三局", "护薪", "纸质"]]
            if all(state == "缺勤/无记录" for state in states):
                status = "— 数据缺失"
            elif len(set(states)) == 1 and "待确认" not in states:
                status = "✔ 一致"
            else:
                status = "⚠ 有差异"
            rows.append({
                "分包/所属企业": company,
                "姓名": base.get("姓名", ""),
                "身份证号": base.get("身份证号", ""),
                "日期": day,
                "三局原始内容": values["三局"],
                "护薪原始内容": values["护薪"],
                "纸质原始内容": values["纸质"],
                "三局识别": states[0],
                "护薪识别": states[1],
                "纸质识别": states[2],
                "日差异状态": status,
            })
    return pd.DataFrame(rows)


def _to_float(v):
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return np.nan


def build_diff_table(df_a, df_b, df_paper, company=''):

    def _extract(df, alias):
        if df.empty:
            return pd.DataFrame(columns=['姓名', '身份证号', alias, '__key__'])
        name_col = _resolve_col(df, _NAME_CANDS)
        id_col = _resolve_col(df, _ID_CANDS)
        days_col = _resolve_col(df, _DAYS_CANDS)
        if name_col is None:
            return pd.DataFrame(columns=['姓名', '身份证号', alias, '__key__'])
        out = pd.DataFrame()
        out['姓名'] = df[name_col].astype(str).str.strip()
        out['身份证号'] = df[id_col].apply(_clean_identity) if id_col else ''
        out[alias] = df[days_col].apply(_to_float) if days_col else np.nan
        out = out[
            out.apply(
                lambda row: _is_valid_identity(row['身份证号']) or _looks_like_person_name(row['姓名']),
                axis=1,
            )
        ]
        return out.reset_index(drop=True)

    ea = _extract(df_a, '三局天数')
    eb = _extract(df_b, '护薪天数')
    ep = _extract(df_paper, '纸质天数')

    # 先收集三份考勤中“姓名 -> 身份证号”的唯一映射；再补充主数据中的同公司唯一姓名。
    # 这使没有身份证号的 6 月纸质模板仍能合并到三局/护薪/主数据里的同一人员。
    name_to_ids = {}

    def _add_name_id(name, identity):
        clean_name = _raw_attendance_value(name)
        clean_id = _clean_identity(identity)
        if clean_name and _is_valid_identity(clean_id):
            name_to_ids.setdefault(clean_name, set()).add(clean_id)

    for part in [ea, eb, ep]:
        for _, row in part.iterrows():
            _add_name_id(row.get('姓名', ''), row.get('身份证号', ''))

    master = _load_master_data()
    if master is not None and not master.empty and {'姓名', '身份证号'}.issubset(master.columns):
        master_scope = master.copy()
        if company and '分包/所属企业' in master_scope.columns:
            scoped = master_scope[master_scope['分包/所属企业'].astype(str).str.strip() == str(company).strip()]
            if not scoped.empty:
                master_scope = scoped
        for _, row in master_scope.iterrows():
            _add_name_id(row.get('姓名', ''), row.get('身份证号', ''))

    for part in [ea, eb, ep]:
        if part.empty:
            continue
        for idx, row in part.iterrows():
            identity = _clean_identity(row.get('身份证号', ''))
            if not _is_valid_identity(identity):
                candidates = name_to_ids.get(_raw_attendance_value(row.get('姓名', '')), set())
                if len(candidates) == 1:
                    identity = next(iter(candidates))
                    part.at[idx, '身份证号'] = identity
            part.at[idx, '__key__'] = identity if _is_valid_identity(identity) else f"姓名:{row.get('姓名', '')}"
        part.drop_duplicates(subset='__key__', keep='first', inplace=True)
        part.reset_index(drop=True, inplace=True)

    merged = ea.copy()
    for part in [eb, ep]:
        if part.empty or '__key__' not in part.columns:
            continue
        merged = merged.merge(part, on='__key__', how='outer', suffixes=('', '_new'))
        for col in ['姓名', '身份证号']:
            new_col = f'{col}_new'
            if new_col in merged.columns:
                merged[col] = merged[col].replace('', pd.NA).fillna(merged[new_col])
                merged.drop(columns=[new_col], inplace=True)

    for col in ['三局天数', '护薪天数', '纸质天数']:
        if col not in merged.columns:
            merged[col] = np.nan

    def _status(row):
        vals = [row['三局天数'], row['护薪天数'], row['纸质天数']]
        valid = [v for v in vals if not (isinstance(v, float) and np.isnan(v))]
        if len(valid) == 0:
            return '— 数据缺失'
        if len(set(valid)) == 1:
            return '✔ 一致'
        return '⚠ 有差异'

    merged['差异状态'] = merged.apply(_status, axis=1)
    merged = merged.reset_index(drop=True)
    result = merged[['姓名', '身份证号', '三局天数', '护薪天数', '纸质天数', '差异状态']]
    if company:
        result.insert(0, '分包/所属企业', company)
    return result

def parse_watermark_attendance(df, company='', period=''):
    """
    单源水印签到表解析。
    基于 _safe_read 的结果，优先通过识别每日考勤计算出勤天数，
    如果未找到每日明细，则尝试读取总天数列。
    """
    if df.empty:
        return __import__('pandas').DataFrame()
    
    days_col = _resolve_col(df, _DAYS_CANDS)
    records = _extract_daily_source(df, '水印签到表')
    
    rows = []
    if records:
        for key, item in records.items():
            name = item.get('姓名', '')
            identity = item.get('身份证号', '')
            days_dict = item.get('days', {})
            
            actual_days = 0
            has_daily_records = len(days_dict) > 0
            
            if has_daily_records:
                for day, val in days_dict.items():
                    state = _attendance_state(val)
                    if state == '有考勤':
                        actual_days += 1
            
            if not has_daily_records and days_col:
                name_col = _resolve_col(df, _NAME_CANDS)
                mask = (df[name_col] == name)
                if identity:
                    id_c = _resolve_col(df, _ID_CANDS)
                    if id_c:
                        mask = mask & (df[id_c].apply(_clean_identity) == identity)
                match = df[mask]
                if not match.empty:
                    val = _to_float(match.iloc[0][days_col])
                    actual_days = val if not __import__('pandas').isna(val) else 0
                    
            rows.append({
                '分包/所属企业': company,
                '姓名': name,
                '身份证号': identity,
                '月份': period,
                '__考勤明细__': dict(days_dict),
                '解析出勤天数': float(actual_days),
                '最终核定天数': float(actual_days)
            })
    else:
        name_col = _resolve_col(df, _NAME_CANDS)
        id_col = _resolve_col(df, _ID_CANDS)
        for _, row in df.iterrows():
            name = _raw_attendance_value(row.get(name_col, ''))
            identity = _clean_identity(row.get(id_col, '')) if id_col else ''
            if not _looks_like_person_name(name) and not _is_valid_identity(identity):
                continue
            actual_days = _to_float(row.get(days_col)) if days_col else 0
            if __import__('pandas').isna(actual_days):
                actual_days = 0
            rows.append({
                '分包/所属企业': company,
                '姓名': name,
                '身份证号': identity,
                '月份': period,
                '__考勤明细__': {},
                '解析出勤天数': float(actual_days),
                '最终核定天数': float(actual_days)
            })

    res_df = __import__('pandas').DataFrame(rows)
    if not res_df.empty:
        res_df['__key__'] = res_df.apply(lambda r: r['身份证号'] if _is_valid_identity(r['身份证号']) else f"姓名:{r['姓名']}", axis=1)
        res_df.drop_duplicates(subset='__key__', keep='first', inplace=True)
        res_df.drop(columns=['__key__'], inplace=True)
    return res_df


# ================================================================
# 导出层
# ================================================================

def _thin_border():
    s = Side(style='thin', color='D1D5DB')
    return Border(top=s, left=s, right=s, bottom=s)


def _cell_style(cell, font=None, fill=None, align=None, border=None):
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border


def _add_title_row(ws, title, n_cols, fill):
    ws.insert_rows(1)
    ws.row_dimensions[1].height = 36
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(bold=True, size=14, color='FFFFFF', name='微软雅黑')
    c.fill      = fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = _thin_border()
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _ensure_data_rows(ws, start_row, count, footer_row):
    """在模板签字/合计区域前扩展数据行，并复制上一行的样式。"""
    available = max(0, footer_row - start_row)
    extra = count - available
    if extra <= 0:
        return

    # openpyxl 的 insert_rows 不会自动移动 merged_cells；先拆分底部签字区，
    # 插入数据行后再按新行号恢复合并，否则新增数据会写入 MergedCell。
    footer_merges = []
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= footer_row:
            footer_merges.append((merged.min_row, merged.min_col, merged.max_row, merged.max_col))
            ws.unmerge_cells(str(merged))

    ws.insert_rows(footer_row, extra)
    source_row = footer_row - 1
    for row_idx in range(footer_row, footer_row + extra):
        ws.row_dimensions[row_idx].height = ws.row_dimensions[source_row].height
        for col_idx in range(1, ws.max_column + 1):
            source = ws.cell(source_row, col_idx)
            target = ws.cell(row_idx, col_idx)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)

    for min_row, min_col, max_row, max_col in footer_merges:
        ws.merge_cells(
            start_row=min_row + extra,
            start_column=min_col,
            end_row=max_row + extra,
            end_column=max_col,
        )


def _find_footer_row(ws, start_row):
    """识别模板底部声明/签字区域的起始行，兼容旧版和新版工资模板。"""
    keywords = ('申明', '声明', '班组长签字', '第  1  页', '第1页')
    for row_idx in range(start_row, ws.max_row + 1):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, min(ws.max_column, 5) + 1)]
        text = ' '.join(str(value) for value in values if value is not None)
        if any(keyword in text for keyword in keywords):
            return row_idx
    return ws.max_row + 1


def _select_company_sheet(wb, company):
    """按公司简称选择模板 Sheet；没有匹配时才回退到 active。"""
    aliases = []
    if '江苏旭之升' in str(company):
        aliases = ['旭之升', '江苏']
    elif '青海久昌' in str(company):
        aliases = ['青海久昌', '久昌']
    for ws in wb.worksheets:
        if any(alias in ws.title for alias in aliases):
            return ws
    return wb.active


def _keep_only_sheet(wb, ws):
    """公司级导出只保留当前公司的 Sheet，避免打开文件时落到另一家公司空白页。"""
    for other in list(wb.worksheets):
        if other is not ws:
            wb.remove(other)
    wb.active = 0
    return ws


def build_diff_report_xlsx(diff_df):
    if not OPENPYXL_OK:
        return b''
    wb = Workbook()
    ws = wb.active
    ws.title = '三方考勤差异报告'
    ws.freeze_panes = 'A3'

    fill_title  = PatternFill('solid', fgColor='7C3AED')
    fill_header = PatternFill('solid', fgColor='5B21B6')
    fill_ok     = PatternFill('solid', fgColor='D1FAE5')
    fill_warn   = PatternFill('solid', fgColor='FEF3C7')
    fill_miss   = PatternFill('solid', fgColor='F1F5F9')
    fh  = Font(bold=True, color='FFFFFF', size=11, name='微软雅黑')
    fb  = Font(size=10, color='334155', name='微软雅黑')
    fw  = Font(size=10, color='92400E', name='微软雅黑', bold=True)
    fm  = Font(size=10, color='94A3B8', name='微软雅黑', italic=True)
    ac  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bd  = _thin_border()
    today = datetime.now().strftime('%Y年%m月')

    headers = ['序号', '公司', '姓名', '身份证号', '三局系统天数', '护薪系统天数', '纸质天数', '差异状态', '领导批示（手填）']
    ws.row_dimensions[1].height = 30
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        _cell_style(c, font=fh, fill=fill_header, align=ac, border=bd)

    for i, row in diff_df.iterrows():
        r = i + 2
        ws.row_dimensions[r].height = 24
        status = str(row.get('差异状态', ''))
        if status == '✔ 一致':
            row_fill, row_font = fill_ok, fb
        elif status.startswith('⚠'):
            row_fill, row_font = fill_warn, fw
        else:
            row_fill, row_font = fill_miss, fm

        def _clean(k):
            v = row.get(k)
            return '' if (v is None or (isinstance(v, float) and np.isnan(v))) else v

        vals = [
            i + 1,
            row.get('分包/所属企业', ''),
            row.get('姓名', ''),
            _clean('身份证号'),
            _clean('三局天数'),
            _clean('护薪天数'),
            _clean('纸质天数'),
            status,
            '',
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            _cell_style(c, font=row_font, fill=row_fill, align=ac, border=bd)

    _set_col_widths(ws, [6, 22, 12, 20, 11, 11, 11, 14, 24])
    _add_title_row(ws, f'三方考勤差异核查报告 — {today}（仅供领导确认，请勿直接发放）', len(headers), fill_title)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load_master_data():
    try:
        df = load_master_df()
        if df.empty:
            return None
        for col in df.columns:
            df[col] = df[col].fillna('').astype(str).str.strip()
        
        def extract_wage(val):
            if pd.isna(val): return None
            match = re.search(r'(\d+)', str(val))
            return float(match.group(1)) if match else None

        if '结算单价/标准' in df.columns:
            df['提取日薪'] = df['结算单价/标准'].apply(extract_wage)
        else:
            df['提取日薪'] = None
            
        if '身份证号' in df.columns:
            df['身份证号'] = df['身份证号'].astype(str).str.strip()
            df = df.sort_values('身份证号').drop_duplicates(subset=['身份证号'], keep='last')
        else:
            df = df.drop_duplicates(subset=['姓名'])
        return df
    except Exception:
        return None


def enrich_with_master(final_df):
    """按身份证号补齐主数据；身份证号缺失时仅对唯一姓名做安全兜底。"""
    master = _load_master_data()
    if final_df is None or final_df.empty or master is None or master.empty:
        return final_df.copy() if isinstance(final_df, pd.DataFrame) else pd.DataFrame()

    out = final_df.copy()
    for col in ['身份证号', '姓名']:
        if col in out.columns:
            out[col] = out[col].fillna('').astype(str).str.strip()
    master = master.copy()
    master['身份证号'] = master.get('身份证号', '').fillna('').astype(str).str.strip()
    master['姓名'] = master.get('姓名', '').fillna('').astype(str).str.strip()

    master_cols = [c for c in [
        '姓名', '性别', '身份证号', '手机号', '班组', '工种', '人员类型',
        '分包/所属企业', '工资卡号', '开户银行', '提取日薪', '项目全称'
    ] if c in master.columns]
    lookup = master[master_cols].copy()
    lookup = lookup.rename(columns={c: f'主表_{c}' for c in master_cols if c not in ['身份证号']})
    out = out.merge(lookup, on='身份证号', how='left')

    # 主表是最终权威来源；考勤文件只用于三方天数和异常核对。
    for col in ['姓名', '性别', '手机号', '班组', '工种', '分包/所属企业', '工资卡号', '开户银行', '提取日薪', '项目全称']:
        master_col = f'主表_{col}'
        if master_col in out.columns:
            if col not in out.columns:
                out[col] = ''
            out[col] = out[col].replace('', pd.NA).fillna(out[master_col])
            out.drop(columns=[master_col], inplace=True)

    # 只对身份证号确实无法匹配、且主表姓名唯一的人员进行姓名兜底。
    if '姓名' in out.columns and '姓名' in master.columns:
        unique_name = master[master['姓名'].ne('')].groupby('姓名').filter(lambda x: len(x) == 1)
        name_map = unique_name.set_index('姓名').to_dict('index')
        for idx, row in out.iterrows():
            if row.get('身份证号', '') in master['身份证号'].values:
                continue
            candidate = name_map.get(str(row.get('姓名', '')).strip())
            if candidate:
                for col in ['性别', '手机号', '班组', '工种', '分包/所属企业', '工资卡号', '开户银行', '提取日薪', '项目全称']:
                    if (not str(row.get(col, '')).strip()) and candidate.get(col):
                        out.at[idx, col] = candidate[col]

    # 统一导出字段别名，避免模板导出阶段因界面字段名不同而漏填电话/卡号。
    if '联系电话' not in out.columns:
        out['联系电话'] = ''
    if '手机号' in out.columns:
        out['联系电话'] = out['联系电话'].replace('', pd.NA).fillna(out['手机号'])
    if '银行卡号' not in out.columns:
        out['银行卡号'] = ''
    if '工资卡号' in out.columns:
        out['银行卡号'] = out['银行卡号'].replace('', pd.NA).fillna(out['工资卡号'])

    return out


import zipfile
from openpyxl import load_workbook

def _excel_value(row, *names):
    """按候选字段读取值；身份证、卡号等字段由调用方再按文本写入。"""
    for name in names:
        value = row.get(name, '')
        if value is None:
            continue
        try:
            if pd.isna(value):
                continue
        except (TypeError, ValueError):
            pass
        return value
    return ''


def _excel_text(value):
    if value is None:
        return ''
    try:
        if pd.isna(value):
            return ''
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _period_label(value):
    text = _excel_text(value)
    if not text:
        return ''
    match = re.search(r'(\d{1,2})\s*月', text)
    if match:
        return f'{int(match.group(1))}月'
    match = re.search(r'(?<!\d)(\d{1,2})(?!\d)', text)
    return f'{int(match.group(1))}月' if match else text


def _period_number(value):
    match = re.search(r'(\d{1,2})\s*月', _excel_text(value))
    if match:
        return int(match.group(1))
    return None


def _infer_period(value):
    """从上传文件名或标题提取月份，缺失时保留空值而不擅自改变业务月份。"""
    match = re.search(r'(\d{1,2})\s*月', _excel_text(value))
    return f'{int(match.group(1))}月' if match else ''


def _short_company(company, attendance=False):
    text = _excel_text(company)
    if '江苏旭之升' in text:
        return '旭之升' if attendance else '江苏旭之升'
    if '青海久昌' in text:
        return '青海久昌'
    return text or '未知公司'


def _first_nonempty(df, columns, fallback=''):
    for column in columns:
        if column not in df.columns:
            continue
        values = df[column].dropna().map(_excel_text)
        values = values[values.ne('')]
        if not values.empty:
            return values.iloc[0]
    return fallback


def _export_batches(salary_df):
    """按公司+月份组织批次；一个批次最终对应一个明细 sheet。"""
    if salary_df is None or salary_df.empty:
        return []
    data = salary_df.copy()
    if '分包/所属企业' not in data.columns:
        data['分包/所属企业'] = '未知公司'
    if '月份' not in data.columns:
        data['月份'] = ''
    data['分包/所属企业'] = data['分包/所属企业'].map(_excel_text).replace('', '未知公司')
    data['月份'] = data['月份'].map(_period_label)
    batches = []
    for (company, period), group in data.groupby(['分包/所属企业', '月份'], sort=False, dropna=False):
        batches.append({
            'company': _excel_text(company) or '未知公司',
            'period': _period_label(period),
            'df': group.reset_index(drop=True),
        })
    return batches


def _safe_sheet_title(title, used_titles):
    title = re.sub(r'[\\/*?:\[\]]', '', _excel_text(title)) or '明细'
    title = title[:31]
    base = title
    suffix = 2
    while title in used_titles:
        tail = f'_{suffix}'
        title = f'{base[:31-len(tail)]}{tail}'
        suffix += 1
    used_titles.add(title)
    return title


def _copy_template_sheets(template_path, batches, name_builder, company_specific=True):
    wb = load_workbook(template_path)
    source_pairs = []
    for batch in batches:
        source = _select_company_sheet(wb, batch['company']) if company_specific else wb.active
        source_pairs.append((batch, source))
    # 先把模板原 sheet 改成临时名，避免 openpyxl 在复制后因同名自动追加“1”。
    for index, source in enumerate(list(wb.worksheets)):
        source.title = f'__模板_{index}'
    outputs = []
    used_titles = set()
    for batch, source in source_pairs:
        sheet = wb.copy_worksheet(source)
        sheet.title = _safe_sheet_title(name_builder(batch), used_titles)
        outputs.append((batch, sheet))
    output_sheets = {sheet for _, sheet in outputs}
    for sheet in list(wb.worksheets):
        if sheet not in output_sheets:
            wb.remove(sheet)
    if wb.worksheets:
        wb.active = 0
    return wb, outputs


def _clear_rows(ws, start_row, end_row, start_col=1, end_col=None):
    if end_row < start_row:
        return
    end_col = end_col or ws.max_column
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                cell.value = None
            except AttributeError:
                pass


def _find_attendance_footer_row(ws, start_row):
    keywords = ('分包项目负责人', '分包班组长', '制表人', '总包负责人', '申明', '声明', '第  1  页', '第1页')
    for row_idx in range(start_row, ws.max_row + 1):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        text = ' '.join(_excel_text(value) for value in values if value is not None)
        if any(keyword in text for keyword in keywords):
            return row_idx
    return ws.max_row + 1


def _prepare_data_area(ws, start_row, count, footer_finder):
    footer_row = footer_finder(ws, start_row)
    _ensure_data_rows(ws, start_row, count, footer_row=footer_row)
    footer_row = footer_finder(ws, start_row)
    _clear_rows(ws, start_row, footer_row - 1)
    return footer_row


def _header_row(ws, required='姓名'):
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        labels = [_normalize_header(ws.cell(row_idx, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
        if any(required in label for label in labels):
            return row_idx
    return 1


def _header_col(ws, row_idx, labels):
    normalized_labels = [_normalize_header(label) for label in labels]
    for col_idx in range(1, ws.max_column + 1):
        current = _normalize_header(ws.cell(row_idx, col_idx).value)
        if any(label == current or label in current for label in normalized_labels):
            return col_idx
    return None


def _write_text_cell(ws, row_idx, col_idx, value):
    if not col_idx:
        return
    cell = ws.cell(row=row_idx, column=col_idx)
    cell.value = _excel_text(value)
    cell.number_format = '@'


def _write_number_cell(ws, row_idx, col_idx, value):
    if not col_idx:
        return
    text = _excel_text(value)
    if not text:
        ws.cell(row=row_idx, column=col_idx).value = None
        return
    number = _to_float(text)
    if pd.isna(number):
        ws.cell(row=row_idx, column=col_idx).value = value
    else:
        ws.cell(row=row_idx, column=col_idx).value = int(number) if float(number).is_integer() else number


def _daily_marks(row):
    marks = _excel_value(row, '__考勤明细__')
    return marks if isinstance(marks, dict) else {}


def _write_daily_cells(ws, row_idx, row, marker, start_col=4, day_count=31):
    marks = _daily_marks(row)
    for day in range(1, day_count + 1):
        value = marks.get(day, marks.get(str(day), ''))
        cell = ws.cell(row=row_idx, column=start_col + day - 1)
        cell.value = marker if _attendance_state(value) == '有考勤' else None


def _batch_project(group):
    return _first_nonempty(group, ['项目全称', '项目简称'], '科技文化中心—国际体育中心（足球场项目）')


def _batch_team(group):
    return _first_nonempty(group, ['班组'], '各班组')


def _set_attendance_sheet_header(ws, batch, standard):
    group = batch['df']
    project = _batch_project(group)
    team = _batch_team(group)
    period = batch['period']
    if not period:
        return
    year = datetime.now().year
    month_number = _period_number(period)
    month_text = f'{year}年{month_number}月' if month_number else period
    if standard == 'zongbao':
        ws['A1'] = f'{project}工程{period}考勤表'
        ws['A2'] = batch['company']
        ws['O2'] = f'班组：{team}'
        ws['AE2'] = f'日期：{month_text}'
    else:
        ws['A4'] = f'项目名称（全称）：{project}                        班组名称：{team}                        {month_text}'


def _normalize_company_attendance_grid(ws):
    """将公司模板的“上月26-31+本月1-25”表头统一为参考表的1-31日列。"""
    header_styles = {}
    for col_idx in range(1, 36):
        source = ws.cell(row=6, column=col_idx)
        header_styles[col_idx] = copy(source._style)
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= 6 and merged.max_row >= 5 and merged.min_col <= 35 and merged.max_col >= 1:
            ws.unmerge_cells(str(merged))
    ws['A5'] = '编号'
    ws['B5'] = '姓名'
    ws['C5'] = '工种'
    for day in range(1, 32):
        cell = ws.cell(row=5, column=3 + day)
        cell.value = day
        if header_styles[3 + day]:
            cell._style = copy(header_styles[3 + day])
    ws['AI5'] = '合计'
    if header_styles[35]:
        ws['AI5']._style = copy(header_styles[35])
    return 6


def _populate_company_attendance_sheet(ws, batch):
    group = batch['df']
    start_row = _normalize_company_attendance_grid(ws)
    footer_row = _prepare_data_area(ws, start_row, len(group), _find_attendance_footer_row)
    _set_attendance_sheet_header(ws, batch, 'company')
    for row_idx, (_, row) in enumerate(group.iterrows(), start=start_row):
        ws.cell(row=row_idx, column=1, value=row_idx - start_row + 1)
        _write_text_cell(ws, row_idx, 2, _excel_value(row, '姓名'))
        _write_text_cell(ws, row_idx, 3, _excel_value(row, '工种'))
        _write_daily_cells(ws, row_idx, row, '✓')
        _write_number_cell(ws, row_idx, 35, _excel_value(row, '最终核定天数'))
    return {'start_row': start_row, 'footer_row': footer_row}


def _populate_zongbao_attendance_sheet(ws, batch):
    group = batch['df']
    start_row = 5
    footer_row = _prepare_data_area(ws, start_row, len(group), _find_attendance_footer_row)
    _set_attendance_sheet_header(ws, batch, 'zongbao')
    for row_idx, (_, row) in enumerate(group.iterrows(), start=start_row):
        ws.cell(row=row_idx, column=1, value=row_idx - start_row + 1)
        _write_text_cell(ws, row_idx, 2, _excel_value(row, '姓名'))
        _write_text_cell(ws, row_idx, 3, _excel_value(row, '性别'))
        _write_daily_cells(ws, row_idx, row, 8)
        _write_number_cell(ws, row_idx, 35, _excel_value(row, '最终核定天数'))
    return {'start_row': start_row, 'footer_row': footer_row}


def _set_total_wage_header(ws, batch):
    period = batch['period']
    if not period:
        return
    ws['A2'] = f'项目名称：{_batch_project(batch["df"])}'
    ws['D2'] = f'劳务单位：{batch["company"]}'
    ws['H2'] = f'月度：{period}'


def _populate_zongbao_wage_sheet(ws, batch):
    group = batch['df']
    start_row = 4
    footer_row = _prepare_data_area(ws, start_row, len(group), lambda sheet, start: _find_footer_row(sheet, start))
    _set_total_wage_header(ws, batch)
    for row_idx, (_, row) in enumerate(group.iterrows(), start=start_row):
        ws.cell(row=row_idx, column=1, value=row_idx - start_row + 1)
        _write_text_cell(ws, row_idx, 2, _excel_value(row, '姓名'))
        _write_text_cell(ws, row_idx, 3, _excel_value(row, '身份证号'))
        _write_text_cell(ws, row_idx, 4, _excel_value(row, '联系电话', '手机号', '电话', '手机号码'))
        _write_text_cell(ws, row_idx, 5, _excel_value(row, '开户银行'))
        _write_text_cell(ws, row_idx, 6, _excel_value(row, '银行卡号', '工资卡号', '卡号'))
        _write_number_cell(ws, row_idx, 7, _excel_value(row, '最终核定天数'))
        _write_number_cell(ws, row_idx, 8, _excel_value(row, '应发工资'))
    ws.column_dimensions['C'].width = max(ws.column_dimensions['C'].width or 0, 24)
    ws.column_dimensions['D'].width = max(ws.column_dimensions['D'].width or 0, 16)
    ws.column_dimensions['F'].width = max(ws.column_dimensions['F'].width or 0, 24)
    return {'start_row': start_row, 'footer_row': footer_row, 'pay_col': 8}


def _set_wage_sheet_header(ws, comp_df, company, period=''):
    """更新新版公司工资模板的合并表头，按导出批次写入月份。"""
    project = _first_nonempty(comp_df, ['项目全称', '项目简称'], '科技文化中心—国际体育中心（足球场项目）')
    team = _first_nonempty(comp_df, ['班组'], '各班组')
    original = str(ws['A2'].value or '')
    month_match = re.search(r'\d{4}\s*年[^\n]*?月', original)
    month_text = month_match.group(0) if month_match else '年  月'
    if period:
        month_number = _period_number(period)
        month_text = f'{datetime.now().year}年{month_number}月' if month_number else period
    ws['A2'] = f'项目名称（全称）：{project}             班组名称：{team}                    {month_text}'


def _populate_company_wage_sheet(ws, batch):
    group = batch['df']
    start_row = 5
    footer_row = _prepare_data_area(ws, start_row, len(group), lambda sheet, start: _find_footer_row(sheet, start))
    _set_wage_sheet_header(ws, group, batch['company'], batch['period'])
    header_row = _header_row(ws, '姓名')
    name_col = _header_col(ws, header_row, ['姓名']) or 2
    job_col = _header_col(ws, header_row, ['工种']) or 3
    days_col = _header_col(ws, header_row, ['出勤工日', '考勤天数']) or 4
    rate_col = _header_col(ws, header_row, ['工资标准', '日薪']) or 5
    gross_col = _header_col(ws, header_row, ['工资总额', '应发工资']) or 6
    pay_col = _header_col(ws, header_row, ['应支付', '实发工资']) or 11
    card_col = _header_col(ws, header_row, ['银行卡号', '工资卡号', '卡号'])
    for row_idx, (_, row) in enumerate(group.iterrows(), start=start_row):
        ws.cell(row=row_idx, column=1, value=row_idx - start_row + 1)
        _write_text_cell(ws, row_idx, name_col, _excel_value(row, '姓名'))
        _write_text_cell(ws, row_idx, job_col, _excel_value(row, '工种'))
        _write_number_cell(ws, row_idx, days_col, _excel_value(row, '最终核定天数'))
        _write_number_cell(ws, row_idx, rate_col, _excel_value(row, '日薪', '提取日薪'))
        _write_number_cell(ws, row_idx, gross_col, _excel_value(row, '应发工资'))
        _write_number_cell(ws, row_idx, pay_col, _excel_value(row, '应发工资'))
        _write_text_cell(ws, row_idx, card_col, _excel_value(row, '银行卡号', '工资卡号', '卡号'))
    return {'start_row': start_row, 'footer_row': footer_row, 'pay_col': pay_col}


def _add_company_wage_summary(wb, details):
    summary = wb.create_sheet('总计', 0)
    summary.merge_cells('A1:E1')
    summary['A1'] = '中建二局安装工程有限公司-工资发放表'
    summary['A1'].font = Font(bold=True, size=16, name='宋体')
    summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
    summary.merge_cells('A2:E2')
    summary['A2'] = '项目名称（全称）：科技文化中心—国际体育中心（足球场项目）'
    summary['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    headers = ['序号', '月份', '劳务分包公司名称', '总额（元）', '合计']
    for col_idx, header in enumerate(headers, start=1):
        summary.cell(row=3, column=col_idx, value=header)
    month_rows = {}
    detail_rows = []
    for idx, detail in enumerate(details, start=1):
        batch = detail['batch']
        row_idx = 3 + idx
        month = _period_number(batch['period'])
        summary.cell(row=row_idx, column=1, value=idx)
        summary.cell(row=row_idx, column=2, value=month)
        summary.cell(row=row_idx, column=3, value=_short_company(batch['company']))
        sheet_name = detail['sheet'].title.replace("'", "''")
        pay_letter = get_column_letter(detail['pay_col'])
        start_row = detail['start_row']
        end_row = start_row + len(batch['df']) - 1
        summary.cell(row=row_idx, column=4, value=f"=SUM('{sheet_name}'!{pay_letter}{start_row}:{pay_letter}{end_row})")
        detail_rows.append(row_idx)
        month_rows.setdefault(month, []).append(row_idx)
    for rows in month_rows.values():
        first = rows[0]
        summary.cell(row=first, column=5, value=f'=SUM(D{rows[0]}:D{rows[-1]})')
        for row_idx in rows[1:]:
            summary.cell(row=row_idx, column=2, value=None)
            summary.cell(row=row_idx, column=5, value=None)
    total_row = 4 + len(detail_rows)
    summary.cell(row=total_row, column=1, value='合计')
    if detail_rows:
        summary.cell(row=total_row, column=4, value=f'=SUM(D{detail_rows[0]}:D{detail_rows[-1]})')
        summary.cell(row=total_row, column=5, value=f'=SUM(E{detail_rows[0]}:E{detail_rows[-1]})')
    summary.column_dimensions['A'].width = 10
    summary.column_dimensions['B'].width = 10
    summary.column_dimensions['C'].width = 24
    summary.column_dimensions['D'].width = 16
    summary.column_dimensions['E'].width = 16
    for row_idx in range(1, total_row + 1):
        for col_idx in range(1, 6):
            summary.cell(row=row_idx, column=col_idx).border = _thin_border()
            summary.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = 'auto'
    except AttributeError:
        pass


def _workbook_bytes(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_all_exports_zip(salary_df):
    """将每个公司/月份批次分别生成四个 Excel 工作簿，并打包为 ZIP。"""
    if not OPENPYXL_OK:
        return b''
    batches = _export_batches(salary_df)
    if not batches:
        return b''
    
    zip_buf = io.BytesIO()
    errors = []
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for batch in batches:
            c_name = _short_company(batch['company'], attendance=False)
            p_name = batch['period'] or "未知月份"
            
            company_str_gongsi = c_name
            period_str_gongsi = p_name
            company_str_zongbao = c_name.replace('-', '_').replace('+', '_')
            period_str_zongbao = p_name.replace('-', '_').replace('+', '_')

            # 公司标准 考勤表
            try:
                wb, outputs = _copy_template_sheets(
                    'load-data/muban/gongsi/附件7：务工人员（含队长、班组长、弄民工）考勤表.xlsx',
                    [batch],
                    lambda b: f"{b['period']}{_short_company(b['company'], attendance=True)}",
                    company_specific=True,
                )
                for b, ws in outputs:
                    _populate_company_attendance_sheet(ws, b)
                filename = f'{company_str_gongsi}-{period_str_gongsi}-考勤表.xlsx'
                zf.writestr(filename, _workbook_bytes(wb))
            except Exception as exc:
                errors.append(f'{c_name} 公司标准考勤表：{exc}')

            # 公司标准 工资确认表
            try:
                wb, outputs = _copy_template_sheets(
                    'load-data/muban/gongsi/副本工资发放表.xlsx',
                    [batch],
                    lambda b: f"{_short_company(b['company'])}{b['period']}",
                    company_specific=True,
                )
                details = []
                for b, ws in outputs:
                    result = _populate_company_wage_sheet(ws, b)
                    details.append({'batch': b, 'sheet': ws, **result})
                _add_company_wage_summary(wb, details)
                filename = f'{company_str_gongsi}-{period_str_gongsi}-工资确认表.xlsx'
                zf.writestr(filename, _workbook_bytes(wb))
            except Exception as exc:
                errors.append(f'{c_name} 公司标准工资确认表：{exc}')

            # 总包标准 考勤表
            try:
                wb, outputs = _copy_template_sheets(
                    'load-data/muban/zongbao/考勤表（一式两份本人签字摁手印）.xlsx',
                    [batch],
                    lambda b: f"{b['period']}{_short_company(b['company'])}",
                    company_specific=False,
                )
                for b, ws in outputs:
                    _populate_zongbao_attendance_sheet(ws, b)
                filename = f'{company_str_zongbao}_{period_str_zongbao}_考勤表.xlsx'
                zf.writestr(filename, _workbook_bytes(wb))
            except Exception as exc:
                errors.append(f'{c_name} 总包标准考勤表：{exc}')

            # 总包标准 工资确认表
            try:
                wb, outputs = _copy_template_sheets(
                    'load-data/muban/zongbao/附件3：农民工资发放确认表(一式两份本人签字摁手印).xlsx',
                    [batch],
                    lambda b: f"{b['period']}{_short_company(b['company'])}",
                    company_specific=False,
                )
                for b, ws in outputs:
                    _populate_zongbao_wage_sheet(ws, b)
                filename = f'{company_str_zongbao}_{period_str_zongbao}_工资确认表.xlsx'
                zf.writestr(filename, _workbook_bytes(wb))
            except Exception as exc:
                errors.append(f'{c_name} 总包标准工资确认表：{exc}')

    for error in errors:
        st.error(f'导出失败：{error}')
    return zip_buf.getvalue()


# ================================================================
# 展示层
# ================================================================

def render():
    st.markdown("""
    <div class="page-header-deco">
        <span class="header-emoji">🗂️</span>
        <div class="header-text">
            <h2>考勤对账与工资结算</h2>
            <p>三方差异可视化 → 领导确认 → 在线定稿 → 一键多模板导出</p>
        </div>
    </div>
    <div class="color-strip"></div>
    """, unsafe_allow_html=True)

    tab_check, tab_export = st.tabs([
        ':material/fact_check: 考勤解析与定稿',
        ':material/payments: 考勤表与工资表导出',
    ])

    with tab_check:
        att_status = st.session_state.get('_att_status', None) # None, 'draft', 'finalized'
        parsed_df = st.session_state.get('_att_parsed_df')
        final_df  = st.session_state.get('final_attendance')

        # ── Step 1：上传与解析 ────────────────────────────
        with st.container(border=True):
            st.markdown("""
            <div class="step-indicator">
                <span class="step-num">1</span>
                <span>上传水印签到表并解析</span>
            </div>
            """, unsafe_allow_html=True)

            st.caption('请上传考勤水印签到表（系统将自动识别 √ 及时间格式计算出勤天数）')
            
            company_names = ['江苏旭之升建筑工程有限公司', '青海久昌建筑装饰工程有限公司']
            company = st.selectbox('默认所属企业（如果文件内未写明公司）', company_names)
            file_watermarks = st.file_uploader('水印签到表 (Excel) - 可多选', type=['xlsx', 'xls'], key='att_watermark', accept_multiple_files=True)

            if st.button(':material/document_scanner: 解析签到表', type='primary', key='btn_parse_watermark'):
                if not file_watermarks:
                    st.warning(':material/info: 请至少上传一份水印签到表')
                else:
                    with st.spinner('正在解析签到表……'):
                        all_res_dfs = []
                        for fw in file_watermarks:
                            df_raw = _safe_read(fw, '水印签到表')
                            period = _infer_period(getattr(fw, 'name', ''))
                            
                            fname = getattr(fw, 'name', '')
                            file_company = company
                            if '旭之升' in fname or '江苏' in fname:
                                file_company = '江苏旭之升建筑工程有限公司'
                            elif '久昌' in fname or '青海' in fname:
                                file_company = '青海久昌建筑装饰工程有限公司'
                            else:
                                try:
                                    if hasattr(fw, 'seek'):
                                        fw.seek(0)
                                    head_df = __import__('pandas').read_excel(fw, header=None, nrows=10)
                                    head_text = head_df.to_string(index=False, header=False)
                                    if '旭之升' in head_text or '江苏' in head_text:
                                        file_company = '江苏旭之升建筑工程有限公司'
                                    elif '久昌' in head_text or '青海' in head_text:
                                        file_company = '青海久昌建筑装饰工程有限公司'
                                except Exception:
                                    pass

                            res_df = parse_watermark_attendance(df_raw, company=file_company, period=period)
                            all_res_dfs.append(res_df)
                        
                        if not all_res_dfs:
                            st.error('解析失败。')
                        else:
                            final_res_df = __import__('pandas').concat(all_res_dfs, ignore_index=True)
                            if final_res_df.empty:
                                st.error('未能识别到有效人员或考勤数据，请检查表格格式。')
                            else:
                                st.session_state['_att_parsed_df'] = final_res_df
                                st.session_state['_att_status'] = 'draft'
                                st.session_state.pop('final_attendance', None)
                                st.session_state.pop('att_data_editor', None)
                                st.success(f'解析成功！共识别到 {len(final_res_df)} 名人员。')
                                st.rerun()
        # ── Step 2：在线确认与定稿 ─────────────────────────────────
        if att_status in ['draft', 'finalized'] and parsed_df is not None and not parsed_df.empty:
            st.markdown('---')
            with st.container(border=True):
                st.markdown("""
                <div class="step-indicator">
                    <span class="step-num">2</span>
                    <span>核对考勤明细与定稿</span>
                </div>
                """, unsafe_allow_html=True)

                if att_status == 'draft':
                    st.markdown(
                        '<div class="hint-box">'
                        ':material/info: 当前为 <b>草稿 (draft)</b> 状态。<br>'
                        '请在下方双击 <b>✏ 最终核定天数</b> 逐行人工确认和微调，确认无误后点击“考勤定稿”。'
                        '</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        '<div class="celebrate-banner" style="margin:0;">'
                        '<span class="celebrate-icon">✅</span>'
                        '考勤已定稿 (finalized)。可进入下一步计算工资或重新修改后再次定稿。'
                        '</div>',
                        unsafe_allow_html=True,
                    )

                # ── 构建基础 DataFrame ─────────────────────────
                edit_base = parsed_df.copy()
                for extra in ['工种', '班组', '联系电话', '开户银行', '银行卡号', '性别', '手机号']:
                    if extra not in edit_base.columns:
                        edit_base[extra] = ''

                # 优先加载已保存的定稿数据
                def _row_key(row):
                    return f"{row.get('分包/所属企业', '')}|{row.get('身份证号', '')}|{row.get('姓名', '')}"

                if final_df is not None and '最终核定天数' in final_df.columns:
                    saved_map = {_row_key(r): r.get('最终核定天数') for _, r in final_df.iterrows()}
                    edit_base['最终核定天数'] = edit_base.apply(
                        lambda r: saved_map.get(_row_key(r), r.get('最终核定天数')), axis=1
                    )

                # ── 在线编辑器 ─────────────────────────────────
                display_edit_cols = [
                    '分包/所属企业', '姓名', '身份证号', '解析出勤天数', '最终核定天数',
                    '工种', '班组', '联系电话', '开户银行', '银行卡号',
                ]
                display_edit_cols = [c for c in display_edit_cols if c in edit_base.columns]

                col_cfg = {
                    '姓名':         st.column_config.TextColumn('姓名', disabled=True),
                    '分包/所属企业': st.column_config.TextColumn('公司', disabled=True),
                    '解析出勤天数': st.column_config.NumberColumn('系统解析天数', disabled=True, format='%.1f'),
                    '最终核定天数': st.column_config.NumberColumn(
                        '✏ 最终核定天数',
                        min_value=0, max_value=31, step=0.5, format='%.1f',
                        help='双击单行进行人工确认/修改',
                    ),
                    '工种':     st.column_config.TextColumn('工种'),
                    '班组':     st.column_config.TextColumn('班组'),
                    '身份证号': st.column_config.TextColumn('身份证号'),
                    '联系电话': st.column_config.TextColumn('联系电话'),
                    '开户银行': st.column_config.TextColumn('开户银行'),
                    '银行卡号': st.column_config.TextColumn('银行卡号'),
                }

                edited = st.data_editor(
                    edit_base[display_edit_cols],
                    column_config=col_cfg,
                    use_container_width=True,
                    num_rows='fixed',
                    hide_index=True,
                    key='att_data_editor',
                )

                c_save, c_info = st.columns([2, 5])
                with c_save:
                    if st.button(':material/check_circle: 考勤定稿', type='primary', key='btn_save_final', use_container_width=True):
                        # data_editor 只返回可见列；把每日明细、月份等隐藏元数据按行补回，
                        # 确保定稿后的工资计算和四类导出仍能追溯原始考勤日格。
                        edited = edited.copy()
                        for hidden_col in edit_base.columns:
                            if hidden_col not in edited.columns:
                                edited[hidden_col] = edit_base[hidden_col].to_numpy()
                        st.session_state['final_attendance'] = enrich_with_master(edited)
                        st.session_state['_att_status'] = 'finalized'
                        st.success(':material/check_circle: 定稿成功 (finalized)！请切换至【考勤表与工资表导出】标签页进行工资计算。')
                        st.rerun()
                with c_info:
                    if att_status == 'finalized':
                        st.markdown('<div class="hint-box" style="margin:0;">:material/info: 已定稿。若重新修改并保存，将更新定稿数据。</div>', unsafe_allow_html=True)

        if att_status is not None:
            with st.expander(':material/refresh: 重新上传 (清空当前数据)'):
                st.warning('将清除当前的草稿和定稿数据，返回初始状态。')
                if st.button(':material/delete_forever: 确认清空', key='btn_reset_tab1', type='secondary'):
                    for k in ['_att_status', '_att_parsed_df', 'final_attendance']:
                        st.session_state.pop(k, None)
                    st.rerun()

    # ============================================================
    # Tab 2 — 导出
    # ============================================================
    with tab_export:
        att_status = st.session_state.get('_att_status')
        final_df = st.session_state.get('final_attendance')

        if att_status != 'finalized' or final_df is None or final_df.empty:
            st.markdown(
                '<div class="alert-box alert-danger" style="margin-top:20px;">'
                ':material/warning: 必须先完成【考勤定稿】(状态: finalized) 才能进行工资计算和导出。'
                '</div>',
                unsafe_allow_html=True,
            )
            return

        with st.container(border=True):
            st.markdown('#### :material/calculate: 工资联动计算')
            wage_mode = st.radio('日薪配置模式', ['使用主表日薪', '全局统一日薪', '按工种分别设置'], horizontal=True, key='wage_mode')
            salary_df = enrich_with_master(final_df)

            if wage_mode == '使用主表日薪':
                salary_df['日薪'] = pd.to_numeric(salary_df.get('提取日薪', 0), errors='coerce')
                missing_wage = salary_df['日薪'].isna() | (salary_df['日薪'] <= 0)
                if missing_wage.any():
                    st.warning(f'有 {int(missing_wage.sum())} 人未能从 data/master 取得有效日薪，请核对主表后再导出。')
            elif wage_mode == '全局统一日薪':
                global_wage = st.number_input('统一日薪（元/天）', min_value=0.0, value=300.0, step=10.0, key='global_wage')
                salary_df['日薪'] = global_wage
            else:
                unique_types = [t for t in final_df.get('工种', pd.Series(dtype=str)).dropna().unique() if str(t).strip()] if '工种' in final_df.columns else []
                if not unique_types:
                    st.info('未检测到工种信息，将使用默认 300 元/天。')
                    salary_df['日薪'] = 300.0
                else:
                    cols = st.columns(min(len(unique_types), 4))
                    wage_map = {}
                    for idx, wt in enumerate(unique_types):
                        with cols[idx % 4]:
                            w = st.number_input(f'「{wt}」日薪', min_value=0.0, value=300.0, step=10.0, key=f'wage_{wt}')
                            wage_map[str(wt).strip()] = w
                    salary_df['日薪'] = salary_df['工种'].apply(lambda x: wage_map.get(str(x).strip(), 300.0))

            def _calc(row):
                try:
                    return round(float(row['最终核定天数']) * float(row['日薪']), 2)
                except (ValueError, TypeError):
                    return ''

            salary_df['应发工资'] = salary_df.apply(_calc, axis=1)

        st.markdown('---')
        st.markdown('#### :material/table_view: 工资预览')
        preview_cols = [c for c in ['姓名','工种','班组','最终核定天数','日薪','应发工资'] if c in salary_df.columns]
        st.dataframe(salary_df[preview_cols], use_container_width=True, hide_index=True)

        valid_salaries = pd.to_numeric(salary_df['应发工资'], errors='coerce').dropna()
        total_pay = valid_salaries.sum()
        avg_days  = pd.to_numeric(salary_df['最终核定天数'], errors='coerce').mean()

        st.markdown(
            f'''
            <div class="metric-container">
                <div class="metric-card"><div class="metric-title">参与结算人数</div><div class="metric-value">{len(salary_df)}</div></div>
                <div class="metric-card"><div class="metric-title">平均核定天数</div><div class="metric-value">{avg_days:.1f} 天</div></div>
                <div class="metric-card"><div class="metric-title">工资总额</div><div class="metric-value">¥{total_pay:,.0f}</div></div>
            </div>
            ''',
            unsafe_allow_html=True,
        )

        st.markdown('---')
        st.markdown('#### :material/download: 按标准模板一键导出')
        st.markdown('<div class="hint-box">系统已读取 `load-data/muban` 目录下的 4 套官方模板，按公司自动分拣，保留所有红头、签字栏及样式进行套打。</div>', unsafe_allow_html=True)

        if not OPENPYXL_OK:
            st.error('未安装 openpyxl，请运行 pip install openpyxl 后重启应用。')
        else:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            
            with st.container(border=True):
                st.markdown('<div style="text-align:center;padding:12px 0"><span style="font-size:40px">🗂️</span><br><b>全量台账压缩包 (ZIP)</b><br><small style="color:#64748B">内含：公司标准考勤、工资表 + 总包标准考勤、工资表<br>已按所属分公司分类整理</small></div>', unsafe_allow_html=True)
                
                zip_data = build_all_exports_zip(salary_df)
                if zip_data:
                    st.download_button(
                        label=':material/archive: 一键打包下载全部 4 套台账',
                        data=zip_data,
                        file_name=f'劳务考勤发薪全套台账_{timestamp}.zip',
                        mime='application/zip',
                        use_container_width=True,
                        type='primary'
                    )
                else:
                    st.error("生成打包文件失败，请检查模板文件是否存在。")
