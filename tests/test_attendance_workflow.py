import pandas as pd
import pytest
from io import BytesIO
from zipfile import ZipFile
from openpyxl import load_workbook

from modules.attendance_payroll import (
    parse_watermark_attendance,
    _attendance_state,
    build_all_exports_zip
)


def test_watermark_attendance_parsing_and_days_calculation():
    """Test that parse_watermark_attendance correctly calculates actual attendance days from daily columns."""
    df_raw = pd.DataFrame([
        {"姓名": "张三", "身份证号": "110101199001011234", "1日": "√", "2日": "√", "3日": "", "4日": "--"},
        {"姓名": "李四", "身份证号": "110101199505054321", "1日": "08:00-18:00", "2日": "请假", "3日": "出勤", "4日": "/"},
    ])
    
    res = parse_watermark_attendance(df_raw, "测试公司")
    
    assert len(res) == 2
    
    # 张三 should have 2 days ("√", "√")
    zhang_san = res[res['姓名'] == '张三'].iloc[0]
    assert zhang_san['解析出勤天数'] == 2.0
    assert zhang_san['最终核定天数'] == 2.0
    assert zhang_san['月份'] == ''
    assert zhang_san['__考勤明细__'][1] == '√'
    assert zhang_san['__考勤明细__'][2] == '√'
    
    # 李四 should have 2 days ("08:00-18:00", "出勤")
    li_si = res[res['姓名'] == '李四'].iloc[0]
    assert li_si['解析出勤天数'] == 2.0


def test_watermark_attendance_fallback_total_days():
    """Test that parse_watermark_attendance falls back to the total days column if daily records are not present."""
    df_raw = pd.DataFrame([
        {"姓名": "王五", "身份证号": "110101199202029876", "考勤天数": "25.5"},
        {"姓名": "赵六", "身份证号": "110101199303031111", "考勤天数": "15"},
    ])
    
    res = parse_watermark_attendance(df_raw, "测试公司")
    
    assert len(res) == 2
    
    wang_wu = res[res['姓名'] == '王五'].iloc[0]
    assert wang_wu['解析出勤天数'] == 25.5
    
    zhao_liu = res[res['姓名'] == '赵六'].iloc[0]
    assert zhao_liu['解析出勤天数'] == 15.0


def test_personnel_identification():
    """Test that personnel identification and deduplication works."""
    df_raw = pd.DataFrame([
        {"姓名": "孙七", "身份证号": "110101199404042222", "1日": "√"},
        {"姓名": "孙七", "身份证号": "110101199404042222", "1日": "√", "2日": "√"},  # Duplicate ID
        {"姓名": "周八", "身份证号": "", "1日": "√"},
        {"姓名": "周八", "身份证号": "", "1日": "√", "2日": "√"},  # Duplicate Name without ID
    ])
    
    res = parse_watermark_attendance(df_raw, "测试公司")
    
    # Should deduplicate based on ID or Name
    assert len(res) == 2
    
    sun_qi = res[res['姓名'] == '孙七'].iloc[0]
    assert sun_qi['解析出勤天数'] == 2.0  # Daily records are merged from both rows


def test_workflow_status_concept():
    """Test conceptual representation of draft and finalized states and payroll gating."""
    # This acts as a mock representation of the UI logic
    
    state = {'_att_status': None, 'final_attendance': None}
    
    # Step 1: Upload and parse -> draft
    state['_att_status'] = 'draft'
    state['_att_parsed_df'] = pd.DataFrame([{"姓名": "张三", "最终核定天数": 20}])
    
    assert state['_att_status'] == 'draft'
    
    # Check if payroll is allowed (simulating the gating logic in Tab 2)
    def can_calculate_payroll(st_state):
        return st_state.get('_att_status') == 'finalized' and st_state.get('final_attendance') is not None
    
    assert not can_calculate_payroll(state)
    
    # Step 2: Finalize
    state['_att_status'] = 'finalized'
    state['final_attendance'] = state['_att_parsed_df']
    
    assert state['_att_status'] == 'finalized'
    assert can_calculate_payroll(state)


def test_export_attendance_and_payroll():
    """Two companies are consolidated into four workbooks with daily marks preserved."""
    salary_df = pd.DataFrame([
        {
            "分包/所属企业": "江苏旭之升建筑工程有限公司",
            "姓名": "张三",
            "身份证号": "110101199001011234",
            "月份": "6月",
            "__考勤明细__": {1: "✓", 2: "✓"},
            "最终核定天数": 20,
            "日薪": 300,
            "应发工资": 6000,
            "工种": "木工",
            "联系电话": "13800000000",
            "开户银行": "工商银行",
            "银行卡号": "6222020000000000000",
            "性别": "男",
        },
        {
            "分包/所属企业": "青海久昌建筑装饰工程有限公司",
            "姓名": "李四",
            "身份证号": "110101199505054321",
            "月份": "6月",
            "__考勤明细__": {1: "8", 3: "出勤"},
            "最终核定天数": 18,
            "日薪": 400,
            "应发工资": 7200,
            "工种": "焊工",
            "联系电话": "13900000000",
            "开户银行": "农业银行",
            "银行卡号": "6222020000000000001",
            "性别": "男",
        },
    ])

    zip_data = build_all_exports_zip(salary_df)

    assert isinstance(zip_data, bytes)
    assert len(zip_data) > 0
    with ZipFile(BytesIO(zip_data)) as archive:
        names = archive.namelist()
        assert len(names) == 4
        assert all(name.endswith('.xlsx') and '/' not in name for name in names)
        assert any('公司标准_考勤表' in name for name in names)
        assert any('公司标准_工资确认表' in name for name in names)
        assert any('总包标准_考勤表' in name for name in names)
        assert any('总包标准_工资确认表' in name for name in names)

        company_attendance = load_workbook(
            BytesIO(archive.read(next(name for name in names if '公司标准_考勤表' in name))),
            data_only=False,
        )
        assert len(company_attendance.sheetnames) == 2
        assert company_attendance[company_attendance.sheetnames[0]]['D6'].value == '✓'
        assert company_attendance[company_attendance.sheetnames[0]]['AI6'].value == 20

        zongbao_attendance = load_workbook(
            BytesIO(archive.read(next(name for name in names if '总包标准_考勤表' in name))),
            data_only=False,
        )
        assert len(zongbao_attendance.sheetnames) == 2
        assert zongbao_attendance[zongbao_attendance.sheetnames[0]]['D5'].value == 8

        company_wage = load_workbook(
            BytesIO(archive.read(next(name for name in names if '公司标准_工资确认表' in name))),
            data_only=False,
        )
        assert company_wage.sheetnames[0] == '总计'
        assert len(company_wage.sheetnames) == 3
        detail = company_wage[company_wage.sheetnames[1]]
        card_col = next(cell.column for cell in detail[3] if cell.value == '银行卡号')
        assert detail.cell(row=5, column=card_col).data_type == 's'
        assert detail.cell(row=5, column=card_col).value == '6222020000000000000'
